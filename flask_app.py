import os
import re
import json
import csv
import io
import base64
import email as email_lib
import email.mime.text
from datetime import datetime
from urllib.parse import quote_plus, urljoin, urlparse

# Load .env so keys work locally without pasting them in the UI every time.
# 1) Walk upward from this file until a .env is found (covers nested layouts)
# 2) Current working directory — fills any vars still missing
# 3) If python-dotenv is missing, parse .env manually (utf-8-sig strips BOM)

_DOTENV_LOADED_PATH = None


def _manual_load_env_file(path: str):
    """Parse KEY=VALUE lines into os.environ (no python-dotenv required)."""
    if not path or not os.path.isfile(path):
        return
    try:
        with open(path, encoding="utf-8-sig") as f:
            for raw in f:
                line = raw.strip()
                if not line or line.startswith("#"):
                    continue
                if "=" not in line:
                    continue
                k, v = line.split("=", 1)
                k = k.strip()
                v = v.strip().strip('"').strip("'")
                if k:
                    os.environ[k] = v
    except OSError:
        pass


def _load_env_files():
    global _DOTENV_LOADED_PATH
    _start_dir = os.path.dirname(os.path.abspath(__file__))
    _env_project = None
    _cur = _start_dir
    for _ in range(10):
        _candidate = os.path.join(_cur, ".env")
        if os.path.isfile(_candidate):
            _env_project = _candidate
            break
        _parent = os.path.dirname(_cur)
        if _parent == _cur:
            break
        _cur = _parent
    _env_cwd = os.path.join(os.getcwd(), ".env")

    def _load(path: str, override: bool):
        if not path or not os.path.isfile(path):
            return
        try:
            from dotenv import load_dotenv

            load_dotenv(path, override=override)
        except ImportError:
            if override:
                _manual_load_env_file(path)
            else:
                # merge: only set keys not already present
                try:
                    with open(path, encoding="utf-8-sig") as f:
                        for raw in f:
                            line = raw.strip()
                            if not line or line.startswith("#") or "=" not in line:
                                continue
                            k, v = line.split("=", 1)
                            k = k.strip()
                            if k and k not in os.environ:
                                v = v.strip().strip('"').strip("'")
                                os.environ[k] = v
                except OSError:
                    pass

    # Prefer project .env, then fill from cwd without overriding
    if _env_project:
        _load(_env_project, override=True)
        _DOTENV_LOADED_PATH = os.path.abspath(_env_project)
    if _env_cwd and os.path.isfile(_env_cwd) and (
        not _env_project
        or os.path.normpath(_env_cwd) != os.path.normpath(_env_project)
    ):
        _load(_env_cwd, override=bool(not _env_project))
        if not _DOTENV_LOADED_PATH:
            _DOTENV_LOADED_PATH = os.path.abspath(_env_cwd)


_load_env_files()

import requests
from flask import Flask, request, session, Response, send_file, jsonify, render_template, redirect
import anthropic

# Gmail OAuth imports
try:
    from google_auth_oauthlib.flow import Flow as GoogleFlow
    from google.oauth2.credentials import Credentials as GoogleCredentials
    from google.auth.transport.requests import Request as GoogleAuthRequest
    from googleapiclient.discovery import build as google_build
    _GMAIL_AVAILABLE = True
except ImportError:
    _GMAIL_AVAILABLE = False

app = Flask(__name__)
# Stable secret in dev keeps Flask sessions valid across restarts (optional: set FLASK_SECRET_KEY in .env)
app.secret_key = (
    os.environ.get("FLASK_SECRET_KEY")
    or os.environ.get("SECRET_KEY")
    or os.urandom(24)
)

# API keys saved from Settings UI (survives server restarts; gitignored)
_LOCAL_API_KEYS_FILE = os.path.join(os.path.dirname(os.path.abspath(__file__)), ".local_api_keys.json")
_LOCAL_SECRET_KEYS = frozenset({
    "anthropic_key",
    "apollo_key",
    "hubspot_token",
    "gemini_key",
    "perplexity_key",
    "hunter_key",
})
_LOCAL_CONFIG_KEYS = (
    "anthropic_key",
    "apollo_key",
    "hubspot_token",
    "gemini_key",
    "perplexity_key",
    "hunter_key",
    "model_provider",
    "claude_model",
    "gemini_model",
    "perplexity_model",
    "crm_path",
)


def _read_local_api_config():
    if not os.path.isfile(_LOCAL_API_KEYS_FILE):
        return {}
    try:
        with open(_LOCAL_API_KEYS_FILE, encoding="utf-8") as f:
            data = json.load(f)
        return data if isinstance(data, dict) else {}
    except Exception:
        return {}


def _write_local_api_config(cfg: dict):
    try:
        with open(_LOCAL_API_KEYS_FILE, "w", encoding="utf-8") as f:
            json.dump(cfg, f, indent=0)
    except OSError:
        pass


def _credential(session_key: str, env_name: str) -> str:
    """Session (from cookie), then disk from Settings save, then process env."""
    try:
        s = (session.get(session_key) or "").strip()
    except RuntimeError:
        s = ""
    if s:
        return s
    disk = (_read_local_api_config().get(session_key) or "").strip()
    if disk:
        return disk
    return (os.getenv(env_name) or "").strip()


def _setting_str(session_key: str, default: str) -> str:
    """Non-secret setting: session, then disk, then default."""
    try:
        s = session.get(session_key)
    except RuntimeError:
        s = None
    if s is not None and str(s).strip() != "":
        return str(s).strip()
    disk = _read_local_api_config().get(session_key)
    if disk is not None and str(disk).strip() != "":
        return str(disk).strip()
    return default

# Module-level storage for leads and outreach (per process)
_leads_store    = []
_outreach_store = []
_perf_store     = []   # performance records [{provider, model, duration_ms, ...}]
_hunter_key     = ""   # Hunter.io API key (set from settings)

# External system safety caps
APOLLO_PULL_LIMIT  = 20
HUBSPOT_PUSH_LIMIT = 20

# External connector safety caps
_APOLLO_MAX_RESULTS = 20
_HUBSPOT_MAX_UPLOAD = 20

# Gmail OAuth storage (persisted to file so server restarts don't lose it)
_GMAIL_TOKEN_FILE   = os.path.join(os.path.dirname(__file__), ".gmail_token.json")
_GMAIL_CLIENT_FILE  = os.path.join(os.path.dirname(__file__), ".gmail_client.json")
_GMAIL_SCOPES       = [
    "https://www.googleapis.com/auth/gmail.send",
    "https://www.googleapis.com/auth/gmail.readonly",
]
_GMAIL_REDIRECT_URI = "http://localhost:8502/api/gmail/callback"

def _load_gmail_creds():
    """Load Gmail OAuth credentials from disk, refresh if expired."""
    if not os.path.exists(_GMAIL_TOKEN_FILE):
        return None
    try:
        with open(_GMAIL_TOKEN_FILE) as f:
            data = json.load(f)
        creds = GoogleCredentials(
            token=data.get("token"),
            refresh_token=data.get("refresh_token"),
            token_uri="https://oauth2.googleapis.com/token",
            client_id=data.get("client_id"),
            client_secret=data.get("client_secret"),
            scopes=_GMAIL_SCOPES,
        )
        if creds.expired and creds.refresh_token:
            creds.refresh(GoogleAuthRequest())
            _save_gmail_creds(creds, data.get("client_id"), data.get("client_secret"))
        return creds
    except Exception:
        return None

def _save_gmail_creds(creds, client_id, client_secret):
    with open(_GMAIL_TOKEN_FILE, "w") as f:
        json.dump({
            "token":         creds.token,
            "refresh_token": creds.refresh_token,
            "client_id":     client_id,
            "client_secret": client_secret,
        }, f)

# Approximate cost per 1M tokens (input, output) in USD
_MODEL_PRICING = {
    # Anthropic
    "claude-opus-4-6":            (5.00,  25.00),
    "claude-sonnet-4-6":          (3.00,  15.00),
    "claude-haiku-4-5":           (1.00,   5.00),
    "claude-opus-4-5":            (5.00,  25.00),
    "claude-sonnet-4-5":          (3.00,  15.00),
    "claude-opus-4-1":            (15.00, 75.00),
    # Gemini
    "gemini-2.5-flash-preview-04-17": (0.075, 0.30),
    "gemini-2.5-pro-preview-05-06":   (1.25,  5.00),
    "gemini-2.0-flash":               (0.075, 0.30),
    "gemini-1.5-flash":               (0.075, 0.30),
    "gemini-1.5-pro":                 (1.25,  5.00),
    "gemini-3-flash-preview":         (0.075, 0.30),
    "gemini-3.1-pro-preview":         (1.25,  5.00),
    "gemini-3.1-flash-lite-preview":  (0.25,  1.50),
    # Perplexity
    "sonar-pro":           (3.00, 15.00),
    "sonar":               (1.00,  1.00),
    "sonar-reasoning-pro": (2.00,  8.00),
    "sonar-reasoning":     (1.00,  5.00),
}

def _estimate_cost(model: str, input_tokens: int, output_tokens: int) -> float:
    """Return estimated USD cost for a single request."""
    price_in, price_out = _MODEL_PRICING.get(model, (1.00, 5.00))
    return round((input_tokens * price_in + output_tokens * price_out) / 1_000_000, 6)

def _record_perf(provider: str, model: str, duration_ms: int,
                 input_tokens: int, output_tokens: int,
                 tool_calls: int, leads_found: int, success: bool):
    global _perf_store
    _perf_store.append({
        "ts":            __import__("time").time(),
        "provider":      provider,
        "model":         model,
        "duration_ms":   duration_ms,
        "input_tokens":  input_tokens,
        "output_tokens": output_tokens,
        "tool_calls":    tool_calls,
        "leads_found":   leads_found,
        "success":       success,
        "cost_usd":      _estimate_cost(model, input_tokens, output_tokens),
    })
    # Keep last 200 records
    if len(_perf_store) > 200:
        _perf_store = _perf_store[-200:]

SCRAPE_HEADERS = {
    "User-Agent": (
        "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) "
        "AppleWebKit/537.36 (KHTML, like Gecko) "
        "Chrome/120.0.0.0 Safari/537.36"
    ),
    "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8",
    "Accept-Language": "en-US,en;q=0.9",
}

# ── Lead CSV fields ───────────────────────────────────────────────────────────

LEAD_FIELDS = [
    # Business identity
    "trade_name",
    "entity_name",
    "formation_date",
    "years_in_business",
    "sunbiz_status",
    "sunbiz_url",
    # Business contact
    "general_email",
    "business_phone",
    "address",
    "website",
    # Owner
    "owner_name",
    "owner_email",
    "owner_phone",
    # Registered Agent
    "registered_agent",
    "reg_agent_address",
    "reg_agent_email",
    "reg_agent_phone",
    # Social / reviews
    "instagram_url",
    "facebook_url",
    "google_review_count",
    "google_rating",
    # Extra
    "industry",
    "employees",
    "linkedin_url",
]


def _leads_db_xlsx_path():
    """Absolute path to the MMG business leads Excel database (session > saved settings > env > defaults)."""
    # Session is only available inside a request context
    try:
        from flask import session as _session, has_request_context
        session_path = (_session.get("crm_path") or "").strip() if has_request_context() else ""
    except Exception:
        session_path = ""

    disk_path = (_read_local_api_config().get("crm_path") or "").strip()
    explicit = session_path or disk_path or (os.environ.get("TENANT_CRM_XLSX_PATH") or "").strip()
    if explicit:
        return os.path.expanduser(explicit)
    # Auto-discover common drop locations so the file works without any setup
    candidates = [
        os.path.expanduser("~/Downloads/MMG_Tenant_CRM.xlsx"),
        os.path.expanduser("~/Downloads/MMG Tenant CRM.xlsx"),
        os.path.expanduser("~/Desktop/MMG_Tenant_CRM.xlsx"),
        os.path.expanduser("~/Desktop/MMG Tenant CRM.xlsx"),
        os.path.join(os.path.dirname(os.path.abspath(__file__)), "MMG_Tenant_CRM.xlsx"),
        os.path.join(os.path.dirname(os.path.abspath(__file__)), "MMG Tenant CRM.xlsx"),
    ]
    for p in candidates:
        if os.path.isfile(p):
            return p
    return ""


# Persistent DuckDB - rebuilt only when Excel is newer than the .db file
_tenant_crm_duck = {
    "con":          None,
    "path":         None,   # xlsx path used for current con
    "db_path":      None,   # .db file path
    "mtime":        None,   # xlsx mtime at last build
    "sheet":        None,
    "table":        "leads_db",
    "row_count":    0,
    "identifiers": [],
}


def _leads_db_path(xlsx_path):
    """Return the .db file path next to the Excel file."""
    base = os.path.splitext(xlsx_path)[0]
    return base + "_leads.duckdb"


def _tenant_crm_cell_str(v):
    if v is None:
        return ""
    if hasattr(v, "isoformat"):
        return v.isoformat()
    return str(v).strip()


def _tenant_crm_unique_sql_idents(headers):
    """Map header labels to unique double-quoted DuckDB identifiers."""
    seen = {}
    out = []
    for i, h in enumerate(headers):
        base = h if h.strip() else f"column_{i + 1}"
        slug = re.sub(r"[^\w]+", "_", base.strip()).strip("_") or f"col_{i + 1}"
        key = slug.lower()
        if key in seen:
            seen[key] += 1
            slug = f"{slug}_{seen[key]}"
        else:
            seen[key] = 0
        ident = '"' + slug.replace('"', '""') + '"'
        out.append(ident)
    return out


def _tenant_crm_load_into_duckdb(path, sheet_name=None):
    """
    Read ALL sheets from .xlsx into a single persistent DuckDB table with a
    'sheet' column. The .db file is only rebuilt when the Excel mtime is newer.
    Returns (sheets_loaded, total_rows_loaded).
    """
    try:
        from openpyxl import load_workbook
    except ImportError:
        raise RuntimeError("openpyxl is not installed. Run: pip install -r requirements.txt") from None
    try:
        import duckdb
    except ImportError:
        raise RuntimeError("duckdb is not installed. Run: pip install -r requirements.txt") from None

    db_path = _leads_db_path(path)
    xlsx_mtime = os.path.getmtime(path)

    # Close any existing open connection before touching the file
    if _tenant_crm_duck["con"] is not None:
        try:
            _tenant_crm_duck["con"].close()
        except Exception:
            pass
        _tenant_crm_duck["con"] = None

    needs_rebuild = (
        not os.path.isfile(db_path)
        or os.path.getmtime(db_path) < xlsx_mtime
    )

    tbl = _tenant_crm_duck["table"]

    if needs_rebuild:
        wb = load_workbook(path, read_only=True, data_only=True)
        sheet_names = wb.sheetnames

        # Build schema from first sheet's header row (all sheets share the same columns)
        first_ws = wb[sheet_names[0]]
        first_rows = first_ws.iter_rows(values_only=True)
        try:
            header_row = next(first_rows)
        except StopIteration:
            wb.close()
            raise ValueError("First sheet is empty") from None

        labels = []
        for i, h in enumerate(header_row):
            if h is None or (isinstance(h, str) and not h.strip()):
                labels.append(f"column_{i + 1}")
            else:
                labels.append(str(h).strip())

        idents = _tenant_crm_unique_sql_idents(labels)
        # Prepend a 'sheet' column so rows are filterable by industry/county
        all_idents = ['"sheet"'] + idents

        if os.path.isfile(db_path):
            os.remove(db_path)
        con = duckdb.connect(database=db_path)
        cols_sql = '"sheet" VARCHAR, ' + ", ".join(f"{ident} VARCHAR" for ident in idents)
        con.execute(f"CREATE TABLE {tbl} ({cols_sql})")
        placeholders = ", ".join(["?"] * len(all_idents))

        total_rows = 0
        for sname in sheet_names:
            ws = wb[sname]
            rows_iter = ws.iter_rows(values_only=True)
            next(rows_iter, None)  # skip header
            batch = []
            for row in rows_iter:
                vals = [sname]
                nonempty = False
                for j in range(len(idents)):
                    cell = row[j] if j < len(row) else None
                    s = _tenant_crm_cell_str(cell)
                    if s:
                        nonempty = True
                    vals.append(s)
                if nonempty:
                    batch.append(tuple(vals))
            if batch:
                con.executemany(f"INSERT INTO {tbl} VALUES ({placeholders})", batch)
                total_rows += len(batch)

        wb.close()
        con.close()

        row_count = total_rows
        truncated = False
        scanned = total_rows
    else:
        # .db is up to date — read schema from it
        con_tmp = duckdb.connect(database=db_path, read_only=True)
        cols_info = con_tmp.execute(f"PRAGMA table_info('{tbl}')").fetchall()
        con_tmp.close()
        all_idents = ['"' + r[1].replace('"', '""') + '"' for r in cols_info]
        idents = [i for i in all_idents if i != '"sheet"']
        row_count = None
        truncated = False
        scanned = 0

    # Keep idents without 'sheet' for text search; store all_idents for SELECT
    con = duckdb.connect(database=db_path, read_only=False)

    _tenant_crm_duck["con"]         = con
    _tenant_crm_duck["path"]        = path
    _tenant_crm_duck["db_path"]     = db_path
    _tenant_crm_duck["mtime"]       = xlsx_mtime
    _tenant_crm_duck["sheet"]       = "all sheets"
    _tenant_crm_duck["row_count"]   = row_count
    _tenant_crm_duck["identifiers"] = all_idents
    _tenant_crm_duck["truncated"]   = truncated
    return sheet_names if needs_rebuild else [], scanned


def _tenant_crm_get_connection(path, sheet_name=None):
    """
    Return a live DuckDB connection.
    - If in-process connection is current (same path + mtime), reuse it.
    - If .db file exists on disk and Excel hasn't changed, open it directly (fast restart).
    - Otherwise rebuild from Excel.
    """
    import duckdb
    xlsx_mtime = os.path.getmtime(path)
    db_path = _leads_db_path(path)
    cache_sheet = sheet_name or ""

    # Reuse open connection if still valid
    if (
        _tenant_crm_duck["con"] is not None
        and _tenant_crm_duck["path"] == path
        and _tenant_crm_duck["mtime"] == xlsx_mtime
        and (_tenant_crm_duck.get("sheet_requested") or "") == cache_sheet
    ):
        return _tenant_crm_duck["con"]

    # Close stale connection
    if _tenant_crm_duck["con"] is not None:
        try:
            _tenant_crm_duck["con"].close()
        except Exception:
            pass
        _tenant_crm_duck["con"] = None

    _tenant_crm_duck["sheet_requested"] = cache_sheet

    # If .db exists and is up to date, open it and reload schema if needed
    if (
        os.path.isfile(db_path)
        and os.path.getmtime(db_path) >= xlsx_mtime
    ):
        con = duckdb.connect(database=db_path, read_only=False)
        if not _tenant_crm_duck["identifiers"]:
            # Fresh process restart — load schema from existing .db
            tbl = _tenant_crm_duck["table"]
            cols_info = con.execute(f"PRAGMA table_info('{tbl}')").fetchall()
            _tenant_crm_duck["identifiers"] = ['"' + r[1].replace('"', '""') + '"' for r in cols_info]
        _tenant_crm_duck["con"]     = con
        _tenant_crm_duck["path"]    = path
        _tenant_crm_duck["db_path"] = db_path
        _tenant_crm_duck["mtime"]   = xlsx_mtime
        return con

    _tenant_crm_load_into_duckdb(path, sheet_name=sheet_name)
    return _tenant_crm_duck["con"]


_CRM_DEFAULT_LIMIT = 50
_CRM_MAX_LIMIT = 5000
_last_user_message = ""


def _normalize_crm_limit(limit, all_results=False):
    """Normalize CRM limit; supports 'all' via all_results or non-positive limits."""
    if all_results:
        return _CRM_MAX_LIMIT
    try:
        n = int(_CRM_DEFAULT_LIMIT if limit is None else limit)
    except Exception:
        n = _CRM_DEFAULT_LIMIT
    if n <= 0:
        return _CRM_MAX_LIMIT
    return max(1, min(n, _CRM_MAX_LIMIT))


_CRM_ALL_WORDS_RE = re.compile(r"\b(all|every|everything|entire|full|complete)\b", re.I)


def _wants_all_results(text: str) -> bool:
    """Heuristic: treat prompts with 'all/every/entire/complete' as full-result queries."""
    if not text:
        return False
    return bool(_CRM_ALL_WORDS_RE.search(text))


def _strip_all_words(text: str) -> str:
    """Remove quantity words like 'all/full/complete' from search text."""
    if not text:
        return ""
    return re.sub(_CRM_ALL_WORDS_RE, " ", text).strip()


def _canonical_query_text(text: str) -> str:
    """Lowercase, remove punctuation, singularize simple plurals, collapse spaces."""
    if not text:
        return ""
    t = re.sub(r"[^\w\s]", " ", text.lower())
    words = []
    for w in t.split():
        # Basic singularization for common plural user terms (salons->salon, barbers->barber)
        if len(w) > 3 and w.endswith("s"):
            w = w[:-1]
        words.append(w)
    return " ".join(words)


def _extract_structured_filters(text: str):
    """
    Infer structured filters from the natural-language query to avoid
    over-restrictive free-text matching.
    Returns dict with optional keys: county, category.
    """
    t = _canonical_query_text(text)
    county = ""
    if "miami dade" in t or "miamidade" in t:
        county = "miami-dade"
    elif "broward" in t:
        county = "broward"

    category = ""
    if "nail salon" in t or "nail" in t:
        category = "nail salon"
    elif "hair salon" in t or ("hair" in t and "salon" in t):
        category = "hair salon"
    elif "barber" in t or "barbershop" in t:
        category = "barbers"

    return {"county": county, "category": category}


def _normalize_query_tokens(text: str):
    """Lowercase + simple singularization + drop filler words."""
    if not text:
        return []
    stop = {
        "show", "me", "find", "get", "list", "pull", "in", "from", "the",
        "a", "an", "for", "and", "or", "of", "some", "please", "i", "want",
        "county", "counties", "database", "db", "result", "results",
    }
    raw = re.sub(r"[^\w\s]", " ", text.lower()).split()
    out = []
    for w in raw:
        if w in stop:
            continue
        # naive singularization helps salon/salons, barber/barbers, etc.
        if len(w) > 4 and w.endswith("s"):
            w = w[:-1]
        out.append(w)
    return out


def _query_structured_hints(query: str):
    """
    Parse common structured hints from natural language for higher recall:
    - county/city tags via sheet filtering
    - industry tags via sheet filtering
    Returns (sheet_terms, keyword_terms).
    """
    toks = _normalize_query_tokens(query)
    tset = set(toks)

    county_terms = []
    if "miami" in tset or "dade" in tset or "miamidade" in tset:
        county_terms.append("miami-dade")
    if "broward" in tset:
        county_terms.append("broward")

    industry_terms = []
    if "nail" in tset:
        industry_terms.append("nail salon")
    if "hair" in tset:
        industry_terms.append("hair salon")
    if "barber" in tset or "barbershop" in tset:
        industry_terms.append("barber")

    sheet_terms = county_terms + industry_terms
    # remove explicitly structured hints from generic keyword matching
    structured = {"miami", "dade", "miamidade", "broward", "nail", "hair", "barber", "barbershop", "salon"}
    keyword_terms = [w for w in toks if w not in structured]
    return sheet_terms, keyword_terms


def _crm_structured_filters(query: str):
    """
    Parse structured filters from natural language query.
    Returns (county_like, industry_like, remaining_words).
    """
    low = (query or "").lower()
    county = ""
    if "miami-dade" in low or "miami dade" in low:
        county = "miami-dade"
    elif "broward" in low:
        county = "broward"

    industry = ""
    if "nail" in low:
        industry = "nail salon"
    elif "barber" in low:
        industry = "barber"
    elif "hair" in low and ("salon" in low or "salons" in low):
        industry = "hair salon"

    tokens = re.findall(r"\w+", low)
    stop = {
        "all", "every", "everything", "entire", "full", "complete",
        "list", "results", "result", "rows", "records", "record", "entries", "entry",
        "show", "me", "find", "get", "give", "pull", "please",
        "in", "of", "for", "from", "the", "a", "an", "to",
        "miami", "dade", "broward", "county",
    }
    if industry == "nail salon":
        stop.update({"nail", "salon", "salons"})
    elif industry == "barber":
        stop.update({"barber", "barbers", "barbershop", "barbershops"})
    elif industry == "hair salon":
        stop.update({"hair", "salon", "salons"})

    words = [t for t in tokens if t not in stop]
    return county, industry, words


def query_tenant_crm(query="", sheet_name=None, limit=50, all_results=False):
    """
    Query the MMG business leads database.
    Routes to Postgres when DATABASE_URL is set (Render), otherwise local DuckDB.
    """
    import logging
    log = logging.getLogger(__name__)

    # Prefer Postgres when DATABASE_URL is available
    if _pg_database_url():
        log.info("[CRM] routing to Postgres: query=%r", query)
        return _query_pg(query=query, limit=limit, all_results=all_results)

    path = _leads_db_xlsx_path()
    log.info("[CRM] query_tenant_crm called: query=%r path=%r", query, path)
    if not path:
        log.warning("[CRM] No Excel path configured")
        return {
            "error": (
                "Business leads database path not set. Add the Excel file path in Settings → Business Leads Database, "
                "or set TENANT_CRM_XLSX_PATH in your .env file."
            ),
            "rows": [],
        }
    if not os.path.isfile(path):
        log.warning("[CRM] Excel file not found: %s", path)
        return {"error": f"Business leads database file not found: {path}", "rows": []}

    if _wants_all_results(f" {query} "):
        all_results = True
    limit = _normalize_crm_limit(limit, all_results=all_results)
    tbl = _tenant_crm_duck["table"]

    try:
        _tenant_crm_get_connection(path, sheet_name=sheet_name)
    except KeyError as e:
        return {"error": str(e), "rows": []}
    except Exception as e:
        return {"error": str(e), "rows": []}

    con = _tenant_crm_duck["con"]
    idents = _tenant_crm_duck["identifiers"]
    if not idents:
        return {
            "rows":    [],
            "count":   0,
            "message": "No columns in sheet",
            "sheet":   _tenant_crm_duck.get("sheet"),
            "path":    path,
            "engine":  "duckdb",
        }

    quoted_cols = ", ".join(ident + " AS " + ident for ident in idents)
    # Include all columns (including 'sheet') in the full-text search hay
    hay = "lower(concat_ws(' ', " + ", ".join(f"nullif({ident}, '')" for ident in idents) + "))"

    q = (query or "").strip()
    if _wants_all_results(f" {q} "):
        all_results = True
    q = _strip_all_words(q)
    sheet_terms, keywords = _query_structured_hints(q)
    words = [w for w in keywords if w]

    try:
        where_parts = []
        params = []
        if sheet_terms:
            sheet_ident = '"sheet"'
            if sheet_ident in idents:
                for st in sheet_terms:
                    where_parts.append(f"lower({sheet_ident}) LIKE ?")
                    params.append(f"%{st}%")
        for w in words:
            where_parts.append(f"{hay} LIKE ?")
            params.append(f"%{w}%")

        if where_parts:
            where = " AND ".join(where_parts)
            sql = f"SELECT {quoted_cols} FROM {tbl} WHERE {where} LIMIT ?"
            result = con.execute(sql, params + [limit]).fetchall()
            cols = [d[0] for d in con.description]
        else:
            sql = f"SELECT {quoted_cols} FROM {tbl} LIMIT ?"
            result = con.execute(sql, [limit]).fetchall()
            cols = [d[0] for d in con.description]

        rows = [dict(zip(cols, r)) for r in result]
        leads = [_normalize_db_lead_row(r) for r in rows]
        global _leads_store
        _leads_store = leads
        log.info("[CRM] query returned %d rows for query=%r", len(rows), query)
        return {
            "rows":    rows,
            "leads":   leads,
            "count":   len(rows),
            "sheet":       _tenant_crm_duck.get("sheet"),
            "source_xlsx": path,
            "db_file":     _tenant_crm_duck.get("db_path"),
            "engine":      "duckdb",
            "table":       tbl,
            "source_rows": _tenant_crm_duck.get("row_count"),
        }
    except Exception as e:
        log.error("[CRM] SQL error: %s", e)
        return {"error": str(e), "rows": [], "engine": "duckdb"}


# ── Postgres leads DB (used on Render / when DATABASE_URL is set) ─────────────

_PG_TABLE   = "leads_db"
_pg_loaded  = False   # True once data has been inserted into Postgres this process


def _lead_row_get(row: dict, *keys, default=""):
    """Fetch first non-empty value among possible key names."""
    for k in keys:
        v = row.get(k)
        if v is not None and str(v).strip():
            return str(v).strip()
    return default


def _extract_city_state(address: str):
    """
    Extract city/state from address patterns like:
    '123 MAIN ST, MIAMI, FL 33101'
    """
    if not address:
        return "", ""
    parts = [p.strip() for p in str(address).split(",") if p.strip()]
    if len(parts) < 2:
        return "", ""
    city = parts[-2] if len(parts) >= 2 else ""
    state = ""
    tail = parts[-1].split()
    if tail:
        state = tail[0]
    return city, state


def _normalize_db_lead_row(row: dict):
    """
    Map DB-native business columns to the lead schema expected by the UI table.
    Keeps a minimal subset required for display and downstream actions.
    """
    address = _lead_row_get(row, "address", "Address", "business_address", "Business_Address")
    city = _lead_row_get(row, "city", "City")
    state = _lead_row_get(row, "state", "State")
    if not city and not state:
        city_guess, state_guess = _extract_city_state(address)
        city = city or city_guess
        state = state or state_guess

    trade_name = _lead_row_get(
        row, "trade_name", "Trade_Name",
        "business_name", "Business_Name",
        "company", "Company",
        "entity_name", "Entity_Name",
    )
    entity_name = _lead_row_get(
        row, "entity_name", "Entity_Name",
        "business_name", "Business_Name",
        "trade_name", "Trade_Name",
    )

    return {
        # Identity
        "trade_name": trade_name,
        "entity_name": entity_name,
        "industry": _lead_row_get(row, "industry", "Industry", "use_category", "Use_Category"),
        # Contact / location
        "business_phone": _lead_row_get(row, "business_phone", "Business_Phone", "phone", "Phone"),
        "general_email": _lead_row_get(row, "general_email", "General_Email", "business_email", "Business_Email", "email", "Email"),
        "address": address,
        "city": city,
        "state": state,
        "website": _lead_row_get(row, "website", "Website", "business_website", "Business_Website"),
        # Social / review fields used by UI (best-effort mappings)
        "instagram_url": _lead_row_get(row, "instagram_url", "Instagram_URL", "instagram_handle", "Instagram_Handle"),
        "facebook_url": _lead_row_get(row, "facebook_url", "Facebook_URL", "facebook_handle", "Facebook_Handle"),
        "google_rating": _lead_row_get(row, "google_rating", "Google_Rating"),
        "google_review_count": _lead_row_get(row, "google_review_count", "Google_Review_Count"),
        # Keep source marker for debugging/filtering
        "sheet": _lead_row_get(row, "sheet", "Sheet"),
    }


def _pg_database_url():
    return (os.environ.get("DATABASE_URL") or "").strip()


def _pg_database_url_ssl():
    """
    Reuse DATABASE_URL, but force sslmode=require when not already present.
    """
    url = _pg_database_url()
    if not url:
        return ""
    if url.startswith("postgres://"):
        url = "postgresql://" + url[len("postgres://"):]
    from urllib.parse import urlsplit, urlunsplit, parse_qsl, urlencode
    parts = urlsplit(url)
    params = dict(parse_qsl(parts.query, keep_blank_values=True))
    if "sslmode" not in params:
        params["sslmode"] = "require"
    query = urlencode(params)
    return urlunsplit((parts.scheme, parts.netloc, parts.path, query, parts.fragment))


def _pg_connect():
    """Return a new psycopg2 connection using DATABASE_URL + sslmode=require."""
    import psycopg2
    url = _pg_database_url_ssl()
    return psycopg2.connect(url)


def _pg_default_seed_csv_path():
    """Fallback seed file for cloud deploys where local Excel path is unavailable."""
    return os.path.join(os.path.dirname(os.path.abspath(__file__)), "leads_data.csv")


def _pg_safe_col(label: str, idx: int) -> str:
    slug = re.sub(r"[^\w]", "_", (label or "")).strip("_").lower()
    return slug or f"col_{idx+1}"


def _pg_rows_from_xlsx(xlsx_path: str):
    """Return (col_names, rows) parsed from all sheets in Excel."""
    from openpyxl import load_workbook

    wb = load_workbook(xlsx_path, read_only=True, data_only=True)
    sheets = wb.sheetnames
    first_iter = wb[sheets[0]].iter_rows(values_only=True)
    header = next(first_iter)
    labels = [str(h).strip() if h and str(h).strip() else f"col_{i+1}" for i, h in enumerate(header)]
    col_names = ["sheet"] + [_pg_safe_col(label, i) for i, label in enumerate(labels)]

    rows = []
    for sname in sheets:
        ws = wb[sname]
        it = ws.iter_rows(values_only=True)
        next(it, None)  # skip header
        for row in it:
            vals = [sname]
            nonempty = False
            for j in range(len(labels)):
                cell = row[j] if j < len(row) else None
                s = _tenant_crm_cell_str(cell)
                if s:
                    nonempty = True
                vals.append(s)
            if nonempty:
                rows.append(tuple(vals))
    wb.close()
    return col_names, rows


def _pg_rows_from_csv(csv_path: str):
    """Return (col_names, rows) parsed from a seed CSV."""
    import csv as _csv

    with open(csv_path, newline="", encoding="utf-8") as f:
        reader = _csv.DictReader(f)
        fields = list(reader.fieldnames or [])
        if not fields:
            raise RuntimeError(f"Seed CSV has no header: {csv_path}")
        col_names = [_pg_safe_col(name, i) for i, name in enumerate(fields)]
        rows = []
        for rec in reader:
            vals = [str(rec.get(name, "") or "").strip() for name in fields]
            if any(vals[1:]):  # keep rows that have non-empty business fields
                rows.append(tuple(vals))
    return col_names, rows


def _pg_ensure_loaded(xlsx_path):
    """
    Create the leads_db table in Postgres if it doesn't exist.
    Load data only if the table is empty (so redeploys don't re-import).
    Source priority:
      1) local Excel path (TENANT_CRM_XLSX_PATH / auto-discovery)
      2) repository seed CSV (leads_data.csv) for Render/cloud
    """
    global _pg_loaded
    import logging
    log = logging.getLogger(__name__)

    con = _pg_connect()
    cur = con.cursor()

    # If table already exists and has data, reuse it.
    cur.execute("SELECT to_regclass(%s)", (_PG_TABLE,))
    exists = cur.fetchone()[0] is not None
    if exists:
        cur.execute(f'SELECT COUNT(*) FROM "{_PG_TABLE}"')
        count = cur.fetchone()[0]
        if count > 0:
            cur.execute("""
                SELECT column_name FROM information_schema.columns
                WHERE table_name = %s ORDER BY ordinal_position
            """, (_PG_TABLE,))
            col_names = [r[0] for r in cur.fetchall()]
            cur.close(); con.close()
            log.info("[PG] leads_db already has %d rows — skipping import", count)
            _pg_loaded = True
            return col_names
        # Existing but empty table; recreate cleanly.
        cur.execute(f'DROP TABLE IF EXISTS "{_PG_TABLE}"')
        con.commit()

    # Choose source rows
    csv_path = _pg_default_seed_csv_path()
    if xlsx_path and os.path.isfile(xlsx_path):
        log.info("[PG] Seeding leads_db from Excel: %s", xlsx_path)
        col_names, rows = _pg_rows_from_xlsx(xlsx_path)
    elif os.path.isfile(csv_path):
        log.info("[PG] Seeding leads_db from CSV: %s", csv_path)
        col_names, rows = _pg_rows_from_csv(csv_path)
    else:
        cur.close(); con.close()
        raise RuntimeError(
            "No seed source available for Postgres. Expected Excel path or leads_data.csv."
        )

    cols_ddl = ", ".join(f'"{c}" TEXT' for c in col_names)
    cur.execute(f'CREATE TABLE "{_PG_TABLE}" ({cols_ddl})')
    con.commit()

    if rows:
        ph = ", ".join(["%s"] * len(col_names))
        cur.executemany(f'INSERT INTO "{_PG_TABLE}" VALUES ({ph})', rows)
        con.commit()
    total = len(rows)

    cur.close(); con.close()
    log.info("[PG] Done — %d total rows loaded", total)
    _pg_loaded = True
    return col_names


def _query_pg(query="", limit=50, all_results=False):
    """Query the Postgres leads_db table. Returns same shape as DuckDB path."""
    import logging
    log = logging.getLogger(__name__)

    xlsx_path = _leads_db_xlsx_path()
    col_names = _pg_ensure_loaded(xlsx_path) if not _pg_loaded else None

    # Fetch column names from DB if we didn't just load
    con = _pg_connect()
    cur = con.cursor()
    if col_names is None:
        cur.execute("""
            SELECT column_name FROM information_schema.columns
            WHERE table_name = %s ORDER BY ordinal_position
        """, (_PG_TABLE,))
        col_names = [r[0] for r in cur.fetchall()]

    if not col_names:
        cur.close(); con.close()
        return {"error": "leads_db table has no columns", "rows": [], "engine": "postgres"}

    cols_sql = ", ".join(f'"{c}"' for c in col_names)
    hay = " || ' ' || ".join(f'COALESCE("{c}", \'\')' for c in col_names)

    q = (query or "").strip()
    if _wants_all_results(f" {q} "):
        all_results = True
    q = _strip_all_words(q)
    words = [w for w in q.lower().split() if w]
    limit = _normalize_crm_limit(limit, all_results=all_results)

    try:
        if not words:
            cur.execute(f'SELECT {cols_sql} FROM "{_PG_TABLE}" LIMIT %s', (limit,))
        else:
            conds = " AND ".join(f"lower({hay}) LIKE %s" for _ in words)
            params = [f"%{w}%" for w in words] + [limit]
            cur.execute(f'SELECT {cols_sql} FROM "{_PG_TABLE}" WHERE {conds} LIMIT %s', params)

        rows = [dict(zip(col_names, r)) for r in cur.fetchall()]
        leads = [_normalize_db_lead_row(r) for r in rows]
        global _leads_store
        _leads_store = leads
        cur.close(); con.close()
        log.info("[PG] query returned %d rows for query=%r", len(rows), query)
        return {
            "rows":    rows,
            "leads":   leads,
            "count":   len(rows),
            "engine":  "postgres",
            "table":   _PG_TABLE,
        }
    except Exception as e:
        cur.close(); con.close()
        log.error("[PG] query error: %s", e)
        return {"error": str(e), "rows": [], "engine": "postgres"}


# ── Tool definitions ──────────────────────────────────────────────────────────

TOOLS = [
    {
        "name": "search_businesses_maps",
        "description": (
            "FALLBACK lead discovery tool — only use when query_tenant_crm returns 0 results. "
            "Searches Google Maps for businesses by keyword and location using a real browser. "
            "Returns trade name, address, city, state, phone, website, Google rating and review count."
        ),
        "input_schema": {
            "type": "object",
            "properties": {
                "keyword":     {"type": "string", "description": "Business type, e.g. 'nail salon'"},
                "location":    {"type": "string", "description": "City and state, e.g. 'Miami, FL'"},
                "num_results": {"type": "integer", "description": "Number of businesses to return (default 10, max 20)", "default": 10},
            },
            "required": ["keyword", "location"],
        },
    },
    {
        "name": "web_search",
        "description": (
            "Search the web for any information — news, business details, market research, "
            "contact info, pricing, reviews, or anything else. Use this whenever you need "
            "current information from the internet."
        ),
        "input_schema": {
            "type": "object",
            "properties": {
                "query": {"type": "string", "description": "The search query"},
            },
            "required": ["query"],
        },
    },
    {
        "name": "apollo_enrich_existing_leads",
        "description": (
            "ENRICH the leads ALREADY in this session with Apollo data (missing email, "
            "website, phone, industry, employees, LinkedIn). Does NOT add new leads — "
            "only fills in blank fields on existing ones. "
            "Use this whenever the user asks to 'find missing emails on Apollo', "
            "'enrich these with Apollo', 'fill the gaps with Apollo', etc."
        ),
        "input_schema": {
            "type": "object",
            "properties": {},
            "required": [],
        },
    },
    {
        "name": "apollo_search_people",
        "description": (
            "BRAND-NEW lead search on Apollo. Use ONLY when the user explicitly asks "
            "for new Apollo leads (not when they want to enrich existing ones). "
            "This REPLACES the current working set with Apollo orgs. "
            "For 'find missing emails / enrich on Apollo', call apollo_enrich_existing_leads instead. "
            "Hard cap: 20 records per request."
        ),
        "input_schema": {
            "type": "object",
            "properties": {
                "keywords":    {"type": "string", "description": "Industry keywords, e.g. 'nail salon'"},
                "locations":   {"type": "array", "items": {"type": "string"},
                                "description": "Locations, e.g. ['Miami, FL']"},
                "num_results": {"type": "integer", "description": "Number of results (max 20)", "default": 20},
            },
            "required": ["keywords"],
        },
    },
    {
        "name": "enrich_leads_batch",
        "description": (
            "Deep research: enrich the current working leads with Sunbiz, website, reviews, and contact fields. "
            "Call when the user wants a full report or missing fields (Sunbiz, etc.) on leads from query_tenant_crm "
            "or search_businesses_maps. Omit `leads` to use session memory. "
            "NEVER call sunbiz_lookup / scrape_website_contact / get_google_reviews individually."
        ),
        "input_schema": {
            "type": "object",
            "properties": {
                "leads": {
                    "type": "array",
                    "description": "Lead objects to enrich; omit to use leads already in this session",
                    "items": {
                        "type": "object",
                        "properties": {
                            "trade_name":          {"type": "string"},
                            "entity_name":         {"type": "string"},
                            "formation_date":      {"type": "string"},
                            "years_in_business":   {"type": "string"},
                            "sunbiz_status":       {"type": "string"},
                            "sunbiz_url":          {"type": "string"},
                            "general_email":       {"type": "string"},
                            "business_phone":      {"type": "string"},
                            "address":             {"type": "string"},
                            "city":                {"type": "string"},
                            "state":               {"type": "string"},
                            "website":             {"type": "string"},
                            "owner_name":          {"type": "string"},
                            "owner_email":         {"type": "string"},
                            "owner_phone":         {"type": "string"},
                            "registered_agent":    {"type": "string"},
                            "reg_agent_address":   {"type": "string"},
                            "reg_agent_email":     {"type": "string"},
                            "reg_agent_phone":     {"type": "string"},
                            "instagram_url":       {"type": "string"},
                            "facebook_url":        {"type": "string"},
                            "google_rating":       {"type": "string"},
                            "google_review_count": {"type": "string"},
                            "industry":            {"type": "string"},
                            "employees":           {"type": "string"},
                            "linkedin_url":        {"type": "string"},
                        },
                    },
                },
            },
            "required": [],
        },
    },
    {
        "name": "hubspot_create_contact",
        "description": (
            "Create or update a HubSpot contact by email. Use after enrich_leads_batch when the user wants "
            "HubSpot records updated with new Sunbiz/phone/website data — call once per lead with the same email."
        ),
        "input_schema": {
            "type": "object",
            "properties": {
                "email":        {"type": "string"},
                "first_name":   {"type": "string"},
                "last_name":    {"type": "string"},
                "company":      {"type": "string"},
                "phone":        {"type": "string"},
                "website":      {"type": "string"},
                "job_title":    {"type": "string"},
                "linkedin":     {"type": "string"},
            },
            "required": ["email"],
        },
    },
    {
        "name": "save_leads_csv",
        "description": (
            "Save the fully enriched lead list to leads.csv. "
            "Call this after enriching leads with Sunbiz, website scraping, and Google reviews."
        ),
        "input_schema": {
            "type": "object",
            "properties": {
                "leads": {
                    "type": "array",
                    "description": "List of enriched lead objects",
                    "items": {
                        "type": "object",
                        "properties": {
                            "trade_name":          {"type": "string"},
                            "entity_name":         {"type": "string"},
                            "formation_date":      {"type": "string"},
                            "years_in_business":   {"type": "string"},
                            "general_email":       {"type": "string"},
                            "owner_name":          {"type": "string"},
                            "owner_email":         {"type": "string"},
                            "owner_phone":         {"type": "string"},
                            "registered_agent":    {"type": "string"},
                            "reg_agent_address":   {"type": "string"},
                            "reg_agent_email":     {"type": "string"},
                            "reg_agent_phone":     {"type": "string"},
                            "business_phone":      {"type": "string"},
                            "address":             {"type": "string"},
                            "website":             {"type": "string"},
                            "instagram_url":       {"type": "string"},
                            "facebook_url":        {"type": "string"},
                            "google_review_count": {"type": "string"},
                            "google_rating":       {"type": "string"},
                            "industry":            {"type": "string"},
                            "employees":           {"type": "string"},
                            "linkedin_url":        {"type": "string"},
                            "sunbiz_url":          {"type": "string"},
                            "sunbiz_status":       {"type": "string"},
                        },
                    },
                }
            },
            "required": ["leads"],
        },
    },
    {
        "name": "get_collected_leads",
        "description": (
            "Return the leads that were already collected and enriched in this session. "
            "Use this when the user wants to view or work with previously found leads."
        ),
        "input_schema": {
            "type": "object",
            "properties": {},
            "required": [],
        },
    },
    {
        "name": "query_tenant_crm",
        "description": (
            "STEP 1 FOR ALL LEAD REQUESTS. MMG's pre-loaded database of 7,302 businesses "
            "(Miami-Dade and Broward counties: hair salons, barbers, nail salons and more). "
            "Always call this BEFORE search_businesses_maps. Pass industry/city/county keywords as `query`. "
            "Empty `query` returns first rows. Multi-word query requires all words to appear in the row. "
            "Only fall back to Maps search if this returns 0 rows."
        ),
        "input_schema": {
            "type": "object",
            "properties": {
                "query": {
                    "type": "string",
                    "description": "Text to find across columns (case-insensitive). Empty = first rows only.",
                },
                "sheet_name": {
                    "type": "string",
                    "description": "Optional Excel sheet name. Omit to use the first sheet.",
                },
                "limit": {
                    "type": "integer",
                    "description": "Rows to return (default 50). Set all_results=true or limit<=0 for all matches.",
                    "default": 50,
                },
                "all_results": {
                    "type": "boolean",
                    "description": "If true, return all matching rows (up to system max).",
                    "default": False,
                },
            },
            "required": [],
        },
    },
    {
        "name": "upload_leads_to_hubspot",
        "description": (
            "Bulk-upload current session leads to HubSpot (max 20 per call). Use when the user asks to push "
            "prospects to HubSpot. For updating existing contacts after enrichment, use hubspot_create_contact per lead."
        ),
        "input_schema": {
            "type": "object",
            "properties": {},
            "required": [],
        },
    },
    {
        "name": "save_research_report",
        "description": (
            "Persist an in-depth research report (markdown) for the current leads so the user "
            "can view it inline and download it as HTML. Call this AFTER you've written the full "
            "report in your reply when the user asks for a 'research report' or 'deep dive'. "
            "Pass the entire markdown body as `report_markdown` and a short `title` like "
            "'Nail Salon — Tenant Prospecting Report — Miami-Dade — April 2026'."
        ),
        "input_schema": {
            "type": "object",
            "properties": {
                "report_markdown": {
                    "type": "string",
                    "description": "The full research report in markdown — sections, tables, bullets included.",
                },
                "title": {
                    "type": "string",
                    "description": "Short report title shown in the download header.",
                },
            },
            "required": ["report_markdown"],
        },
    },
    {
        "name": "save_outreach_csv",
        "description": "Save email drafts to outreach_drafts.csv.",
        "input_schema": {
            "type": "object",
            "properties": {
                "drafts": {
                    "type": "array",
                    "items": {
                        "type": "object",
                        "properties": {
                            "name":         {"type": "string"},
                            "email":        {"type": "string"},
                            "subject_line": {"type": "string"},
                            "email_body":   {"type": "string"},
                        },
                    },
                    "description": "List of email drafts",
                }
            },
            "required": ["drafts"],
        },
    },
    {
        "name": "send_gmail_email",
        "description": (
            "Send an email via the user's connected Gmail account. "
            "Use this when the user explicitly asks to SEND emails (not draft them). "
            "Only works when Gmail is connected via OAuth."
        ),
        "input_schema": {
            "type": "object",
            "properties": {
                "to":      {"type": "string", "description": "Recipient email address"},
                "subject": {"type": "string", "description": "Email subject line"},
                "body":    {"type": "string", "description": "Plain text email body"},
                "send_all_drafts": {
                    "type": "boolean",
                    "description": "If true, send all saved outreach drafts via Gmail.",
                    "default": False,
                },
            },
            "required": [],
        },
    },
    {
        "name": "create_gmail_drafts",
        "description": (
            "Create Gmail drafts from the outreach emails so the user can review and send them manually from Gmail. "
            "Use this by default when the user asks to 'save to Gmail', 'create drafts', or 'push to Gmail'. "
            "Drafts appear in the user's Gmail Drafts folder — nothing is sent automatically."
        ),
        "input_schema": {
            "type": "object",
            "properties": {
                "draft_all": {
                    "type": "boolean",
                    "description": "If true (default), create a Gmail draft for every saved outreach email.",
                    "default": True,
                },
                "to":      {"type": "string", "description": "Recipient for a single draft (only if draft_all is false)"},
                "subject": {"type": "string", "description": "Subject for a single draft"},
                "body":    {"type": "string", "description": "Body for a single draft"},
            },
            "required": [],
        },
    },
]


# ── Tool implementations ───────────────────────────────────────────────────────

def search_businesses_maps(keyword, location, num_results=10):
    """
    Search Google Maps for businesses using a headless browser.
    Returns structured lead data: name, address, city, state, phone,
    website, Google rating, Google review count.
    """
    from playwright.sync_api import sync_playwright
    import time

    num_results = min(int(num_results or 10), 20)
    query = f"{keyword} {location}"
    businesses = []

    try:
        with sync_playwright() as pw:
            browser = pw.chromium.launch(headless=True, args=["--no-sandbox", "--disable-dev-shm-usage", "--disable-gpu"])
            context = browser.new_context(
                user_agent=(
                    "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) "
                    "AppleWebKit/537.36 (KHTML, like Gecko) "
                    "Chrome/120.0.0.0 Safari/537.36"
                ),
                viewport={"width": 1280, "height": 800},
            )
            page = context.new_page()

            search_url = f"https://www.google.com/maps/search/{quote_plus(query)}"
            page.goto(search_url, wait_until="domcontentloaded", timeout=30000)
            time.sleep(3)

            # ── Collect card URLs upfront (before any navigation makes them stale) ──
            # Fetch 2x as buffer — inactive Sunbiz leads get filtered out later
            fetch_count = (num_results * 2) + 5
            cards = page.query_selector_all("a.hfpxzc")
            card_urls = []
            for c in cards[:fetch_count]:
                href = c.get_attribute("href") or ""
                if href and href.startswith("https://"):
                    card_urls.append(href)

            for url in card_urls:
                if len(businesses) >= num_results:
                    break
                try:
                    # Navigate directly by URL — no stale element issues
                    page.goto(url, wait_until="domcontentloaded", timeout=20000)
                    time.sleep(3)

                    html = page.content()

                    # ── Business name ──
                    name_el = page.query_selector("h1.DUwDvf, h1.fontHeadlineLarge")
                    name = name_el.inner_text().strip() if name_el else ""

                    # ── Address ──
                    addr_el = page.query_selector(
                        '[data-item-id="address"] .Io6YTe, '
                        'button[data-item-id="address"] .Io6YTe'
                    )
                    address = addr_el.inner_text().strip() if addr_el else ""

                    # ── Parse city/state from address ──
                    city, state = "", ""
                    if address:
                        parts = address.split(",")
                        if len(parts) >= 2:
                            city = parts[-2].strip()
                        state_zip = parts[-1].strip() if parts else ""
                        state_m = re.search(r'\b([A-Z]{2})\b', state_zip)
                        state = state_m.group(1) if state_m else ""

                    # ── Phone ──
                    phone_el = page.query_selector(
                        '[data-item-id*="phone:tel"] .Io6YTe, '
                        'button[data-tooltip="Copy phone number"] .Io6YTe'
                    )
                    phone = phone_el.inner_text().strip() if phone_el else ""

                    # ── Website ──
                    website_el = page.query_selector(
                        'a[data-item-id="authority"], '
                        'a[aria-label*="website" i]'
                    )
                    website = ""
                    if website_el:
                        href = website_el.get_attribute("href") or ""
                        if href and not href.startswith("https://www.google"):
                            website = href.split("?")[0]

                    # ── Rating + Review count ── multi-strategy extraction
                    rating = ""
                    count  = ""

                    # Strategy 1: aria-label on any element containing "stars"
                    for el in page.query_selector_all('[aria-label]'):
                        try:
                            lbl = el.get_attribute("aria-label") or ""
                            if not rating:
                                rm = re.search(r'(\d[\.,]\d)\s*stars?', lbl, re.IGNORECASE)
                                if rm:
                                    rating = rm.group(1).replace(",", ".")
                                else:
                                    rm2 = re.search(r'^(\d)\s*stars?$', lbl.strip(), re.IGNORECASE)
                                    if rm2:
                                        rating = rm2.group(1)
                            if not count:
                                cm = re.search(r'([\d,]+)\s*reviews?', lbl, re.IGNORECASE)
                                if cm:
                                    count = cm.group(1).replace(",", "")
                            if rating and count:
                                break
                        except Exception:
                            continue

                    # Strategy 2: regex scan on full page HTML
                    if not rating:
                        rm = re.search(r'(\d[\.,]\d)\s*stars?', html, re.IGNORECASE)
                        if rm:
                            rating = rm.group(1).replace(",", ".")
                    if not count:
                        cm = re.search(r'([\d,]+)\s*reviews?', html, re.IGNORECASE)
                        if cm:
                            count = cm.group(1).replace(",", "")

                    # Strategy 3: visible text in known rating container elements
                    if not rating:
                        for sel in ('span.ceNzKf', 'div.F7nice > span', 'span.fontBodyMedium'):
                            try:
                                el = page.query_selector(sel)
                                if el:
                                    txt = el.inner_text().strip()
                                    rm = re.search(r'(\d[\.,]\d)', txt)
                                    if rm:
                                        rating = rm.group(1).replace(",", ".")
                                        break
                            except Exception:
                                continue

                    if name:
                        businesses.append({
                            "trade_name":          name,
                            "entity_name":         "",
                            "formation_date":      "",
                            "years_in_business":   "",
                            "sunbiz_status":       "",
                            "sunbiz_url":          "",
                            "general_email":       "",
                            "business_phone":      phone,
                            "address":             address,
                            "city":                city,
                            "state":               state,
                            "website":             website,
                            "owner_name":          "",
                            "owner_email":         "",
                            "owner_phone":         "",
                            "registered_agent":    "",
                            "reg_agent_address":   "",
                            "reg_agent_email":     "",
                            "reg_agent_phone":     "",
                            "instagram_url":       "",
                            "facebook_url":        "",
                            "google_rating":       rating,
                            "google_review_count": count,
                            "industry":            keyword,
                            "employees":           "",
                            "linkedin_url":        "",
                        })

                except Exception:
                    continue

            browser.close()

    except Exception as e:
        return {"error": str(e), "businesses": []}

    return {"leads": businesses, "total": len(businesses)}


def web_search(query):
    try:
        url = f"https://api.duckduckgo.com/?q={quote_plus(query)}&format=json&no_html=1&skip_disambig=1"
        r = requests.get(url, timeout=10, headers={"User-Agent": "Mozilla/5.0"})
        data = r.json()
        results = []
        if data.get("AbstractText"):
            results.append(f"Summary: {data['AbstractText']}")
            if data.get("AbstractURL"):
                results.append(f"Source: {data['AbstractURL']}")
        for topic in data.get("RelatedTopics", [])[:6]:
            if isinstance(topic, dict) and topic.get("Text"):
                results.append(f"- {topic['Text']}")
                if topic.get("FirstURL"):
                    results.append(f"  {topic['FirstURL']}")
        if data.get("Answer"):
            results.append(f"Answer: {data['Answer']}")
        if not results:
            results.append(f"No direct results. Try: https://www.google.com/search?q={quote_plus(query)}")
        return {"results": "\n".join(results), "query": query}
    except Exception as e:
        return {"error": str(e)}


def _apollo_search_orgs(keywords=None, locations=None, per_page=20, apollo_key=""):
    """Raw Apollo organization search — returns list of org dicts or [] on error."""
    if not apollo_key:
        return []
    payload = {
        "page":                        1,
        "per_page":                    int(per_page or 20),
        "q_organization_keyword_tags": [keywords] if keywords else [],
        "organization_locations":      locations or [],
    }
    try:
        r = requests.post(
            "https://api.apollo.io/v1/organizations/search",
            json=payload,
            headers={
                "Content-Type":  "application/json",
                "Cache-Control": "no-cache",
                "X-Api-Key":     apollo_key,
            },
            timeout=30,
        )
        data = r.json()
        return data.get("organizations") or []
    except Exception:
        return []


def _name_match_score(a: str, b: str) -> float:
    """Score similarity between two business names (higher = better)."""
    if not a or not b:
        return 0.0
    a_words = set(re.sub(r"[^a-z0-9\s]", "", a.lower()).split())
    b_words = set(re.sub(r"[^a-z0-9\s]", "", b.lower()).split())
    noise = {"llc", "inc", "corp", "ltd", "co", "the", "and", "of", "&", "a"}
    a_core = a_words - noise
    b_core = b_words - noise
    if not a_core or not b_core:
        return 0.0
    overlap = len(a_core & b_core)
    return overlap / max(len(a_core), len(b_core))


def apollo_enrich_existing_leads(_apollo_key=None):
    """Enrich the leads ALREADY in session with Apollo data (email, website, phone,
    industry, employees, LinkedIn). Does NOT add new leads — only fills in
    missing fields on existing ones. Use this when the user asks to 'find
    missing emails on Apollo' or 'enrich these with Apollo data'."""
    global _leads_store
    apollo_key = _apollo_key or os.getenv("APOLLO_API_KEY", "")
    if not apollo_key:
        return {"error": "Apollo API key not configured. Please set it in Settings."}
    if not _leads_store:
        return {"error": "No leads in session. Pull leads first (e.g. query_tenant_crm), then ask to enrich them with Apollo."}

    enriched_count = 0
    new_emails = 0
    new_websites = 0
    new_phones = 0
    no_match = []

    for lead in _leads_store:
        name = (lead.get("trade_name") or "").strip()
        if not name:
            continue
        city  = (lead.get("city") or "").strip()
        state = (lead.get("state") or "").strip()
        loc_terms = [t for t in [city, state] if t]
        location_query = ", ".join(loc_terms) if loc_terms else None

        orgs = _apollo_search_orgs(
            keywords=name,
            locations=[location_query] if location_query else None,
            per_page=5,
            apollo_key=apollo_key,
        )

        # Fall back to no-location search if first attempt missed
        if not orgs and location_query:
            orgs = _apollo_search_orgs(
                keywords=name, locations=None, per_page=5, apollo_key=apollo_key
            )

        best = None
        best_score = 0.0
        for org in orgs:
            score = _name_match_score(name, org.get("name") or "")
            if score > best_score:
                best_score = score
                best = org

        if not best or best_score < 0.4:
            no_match.append(name)
            continue

        # Pull useful fields and merge ONLY into blank slots
        website = (best.get("website_url") or best.get("primary_domain") or "").strip()
        if website and not website.startswith("http"):
            website = "https://" + website
        if website and not lead.get("website"):
            lead["website"] = website.rstrip("/")
            new_websites += 1

        # Email via Hunter on the discovered domain
        if website and not lead.get("general_email"):
            try:
                em = _hunter_domain_search(website)
                if em:
                    lead["general_email"] = em
                    new_emails += 1
            except Exception:
                pass

        # Phone
        phone = best.get("phone") or ""
        if not phone:
            pp = best.get("primary_phone") or {}
            phone = pp.get("sanitized_number") or pp.get("number") or ""
        if phone and not lead.get("business_phone"):
            lead["business_phone"] = phone
            new_phones += 1

        # Other useful Apollo metadata (only fill if blank)
        if not lead.get("industry") and best.get("industry"):
            lead["industry"] = best.get("industry")
        if not lead.get("employees") and best.get("estimated_num_employees"):
            lead["employees"] = str(best["estimated_num_employees"])
        if not lead.get("linkedin_url") and best.get("linkedin_url"):
            lead["linkedin_url"] = best["linkedin_url"]
        if not lead.get("facebook_url") and best.get("facebook_url"):
            lead["facebook_url"] = best["facebook_url"]
        if not lead.get("address") and best.get("raw_address"):
            lead["address"] = best["raw_address"]

        enriched_count += 1

    _save_leads_to_file(_leads_store)

    return {
        "leads": _leads_store,
        "summary": {
            "total_leads":      len(_leads_store),
            "matched_on_apollo": enriched_count,
            "no_apollo_match":   len(no_match),
            "emails_added":      new_emails,
            "websites_added":    new_websites,
            "phones_added":      new_phones,
        },
        "no_match": no_match,
    }


def apollo_search_people(keywords=None, locations=None, num_results=20, _apollo_key=None):
    """NEW LEAD SEARCH on Apollo. Use only when the user explicitly wants brand-new
    Apollo leads (not when they ask to enrich existing leads). Returns the leads
    and stores them as the working set, REPLACING anything previously in session."""
    global _leads_store
    apollo_key = _apollo_key or os.getenv("APOLLO_API_KEY", "")
    if not apollo_key:
        return {"error": "Apollo API key not configured. Please set it in Settings."}
    requested = int(num_results or 20)
    effective = min(requested, APOLLO_PULL_LIMIT)
    payload = {
        "page":                        1,
        "per_page":                    effective,
        "q_organization_keyword_tags": [keywords] if keywords else [],
        "organization_locations":      locations or [],
    }
    try:
        r = requests.post(
            "https://api.apollo.io/v1/organizations/search",
            json=payload,
            headers={
                "Content-Type":  "application/json",
                "Cache-Control": "no-cache",
                "X-Api-Key":     apollo_key,
            },
            timeout=30,
        )
        data = r.json()
        if "organizations" not in data:
            return {"error": f"Apollo error (HTTP {r.status_code}): {data}"}

        leads = []
        for org in data["organizations"]:
            # Phone — try multiple fields
            phone = org.get("phone") or ""
            if not phone:
                pp = org.get("primary_phone") or {}
                phone = pp.get("sanitized_number") or pp.get("number") or ""

            # Facebook URL — Apollo sometimes returns this
            fb_url = org.get("facebook_url") or ""

            # Founded year → formation date approximation
            founded_year = org.get("founded_year") or ""
            formation_date = f"01/01/{founded_year}" if founded_year else ""
            years_in_business = ""
            if founded_year:
                try:
                    years_in_business = str(datetime.now().year - int(founded_year))
                except Exception:
                    pass

            website = org.get("website_url", "")

            # Pull email via Hunter.io silently
            general_email = _hunter_domain_search(website) if website else ""

            lead = {
                "trade_name":        org.get("name", ""),
                "entity_name":       "",   # filled by sunbiz_lookup
                "formation_date":    formation_date,
                "years_in_business": years_in_business,
                "general_email":     general_email,
                "owner_name":        "",
                "owner_email":       "",
                "owner_phone":       "",
                "registered_agent":  "",   # filled by sunbiz_lookup
                "reg_agent_address": "",
                "business_phone":    phone,
                "address":           org.get("raw_address", ""),
                "website":           website,
                "instagram_url":     "",   # filled by scrape_website_contact
                "facebook_url":      fb_url,
                "google_review_count": "",
                "google_rating":     "",
                "industry":          org.get("industry", ""),
                "employees":         str(org.get("estimated_num_employees", "")),
                "linkedin_url":      org.get("linkedin_url", ""),
                "sunbiz_url":        "",
                "sunbiz_status":     "",
            }
            leads.append(lead)

        _leads_store = leads
        _save_leads_to_file(leads)
        return {
            "leads": leads,
            "total": len(leads),
            "requested": requested,
            "limit_applied": effective,
            "capped": requested > APOLLO_PULL_LIMIT,
        }
    except Exception as e:
        return {"error": str(e)}


def _sunbiz_name_variants(business_name: str):
    """Generate ranked search variants so we don't miss filings registered under slight variations."""
    if not business_name:
        return []
    raw = business_name.strip()
    variants = [raw]

    cleaned = re.sub(r"[\u2018\u2019\u201c\u201d]", "'", raw)
    cleaned = re.sub(r"\s+&\s+", " AND ", cleaned)
    if cleaned != raw:
        variants.append(cleaned)

    no_punct = re.sub(r"[^\w\s'&]", " ", raw)
    no_punct = re.sub(r"\s+", " ", no_punct).strip()
    if no_punct and no_punct not in variants:
        variants.append(no_punct)

    suffix_pat = re.compile(
        r"\s*\b(LLC|L\.L\.C\.|INC|INC\.|CORP|CORPORATION|CO|COMPANY|LTD|LIMITED|"
        r"PA|PLLC|PLC|HOLDINGS?|GROUP|ENTERPRISES?|SERVICES?|ASSOCIATES?)\b\.?\s*$",
        re.IGNORECASE,
    )
    stripped = suffix_pat.sub("", raw).strip()
    if stripped and stripped not in variants:
        variants.append(stripped)

    paren_free = re.sub(r"\s*\([^)]*\)\s*", " ", raw).strip()
    paren_free = re.sub(r"\s+", " ", paren_free)
    if paren_free and paren_free not in variants:
        variants.append(paren_free)

    words = [w for w in re.split(r"\s+", stripped or raw) if w]
    if len(words) >= 3:
        head = " ".join(words[:3])
        if head not in variants:
            variants.append(head)
    if len(words) >= 2:
        head = " ".join(words[:2])
        if head not in variants:
            variants.append(head)

    seen = set()
    out = []
    for v in variants:
        v = v.strip()
        key = v.lower()
        if v and key not in seen:
            seen.add(key)
            out.append(v)
    return out


def sunbiz_lookup(business_name):
    """
    Search Florida Sunbiz corporate registry using a headless browser.
    Tries multiple name variants (suffix-stripped, first 2 words, etc.) so
    near-matches are still returned. Returns entity name, filing date,
    status, registered agent, and owner.
    """
    from playwright.sync_api import sync_playwright
    import time, logging as _logging
    log = _logging.getLogger(__name__)

    variants = _sunbiz_name_variants(business_name)

    try:
        with sync_playwright() as pw:
            browser = pw.chromium.launch(headless=True, args=["--no-sandbox", "--disable-dev-shm-usage", "--disable-gpu"])
            context = browser.new_context(
                user_agent=(
                    "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) "
                    "AppleWebKit/537.36 (KHTML, like Gecko) "
                    "Chrome/120.0.0.0 Safari/537.36"
                ),
                viewport={"width": 1280, "height": 800},
            )
            page = context.new_page()

            def _name_score(search, candidate):
                s_words = set(re.sub(r"[^a-z0-9\s]", "", search.lower()).split())
                c_words = set(re.sub(r"[^a-z0-9\s]", "", candidate.lower().replace("&amp;", "")).split())
                noise = {"llc", "inc", "corp", "ltd", "co", "the", "a", "of", "and", "&"}
                s_core = s_words - noise
                c_core = c_words - noise
                if not s_core:
                    return 0
                exact = len(s_core & c_core)
                partial = sum(1 for sw in s_core for cw in c_core if sw in cw or cw in sw) - exact
                length_penalty = abs(len(s_core) - len(c_core)) * 0.1
                return exact * 2 + partial * 0.5 - length_penalty

            best_href = ""
            best_score = -1
            best_query = ""

            for query in variants:
                try:
                    page.goto(
                        "https://search.sunbiz.org/Inquiry/CorporationSearch/ByName",
                        wait_until="domcontentloaded", timeout=30000,
                    )
                    time.sleep(0.4)
                    page.fill("#SearchTerm", query)
                    page.click("input[type=submit]")
                    page.wait_for_load_state("domcontentloaded", timeout=15000)
                    time.sleep(1.2)
                    html_results = page.content()
                except Exception as e:
                    log.warning("[sunbiz] search failed for %r: %s", query, e)
                    continue

                result_pairs = re.findall(
                    r'href="(/Inquiry/CorporationSearch/SearchResultDetail[^"]+)"[^>]*>\s*([^<]+?)\s*</a>',
                    html_results,
                )
                if not result_pairs:
                    continue

                for href, name in result_pairs:
                    score = _name_score(business_name, name)
                    if score > best_score:
                        best_score = score
                        best_href = href
                        best_query = query

                # If first variant already found a confident match, stop searching
                if best_score >= 2:
                    break

            if not best_href:
                browser.close()
                return {"found": False, "searched": business_name, "tried": variants}

            detail_url = "https://search.sunbiz.org" + best_href.replace("&amp;", "&")
            page.goto(detail_url, wait_until="domcontentloaded", timeout=30000)
            time.sleep(1.4)
            html = page.content()
            browser.close()

        # ── Parse entity type + corporate name ──
        corp_m = re.search(
            r'<div[^>]*class="[^"]*corporationName[^"]*"[^>]*>.*?<p>([^<]+)</p>\s*<p>([^<]+)</p>',
            html, re.DOTALL,
        )
        entity_type = corp_m.group(1).strip() if corp_m else ""
        entity_name = (
            corp_m.group(2).strip().replace("&amp;", "&") if corp_m else ""
        )

        # ── Parse filing info (label → span pairs) ──
        filing = {}
        for label, value in re.findall(
            r'<label[^>]*>\s*([^<]+?)\s*</label>\s*<span>\s*([^<]*?)\s*</span>', html
        ):
            filing[label.strip()] = value.strip()

        date_filed    = filing.get("Date Filed", "")
        status        = filing.get("Status", "")
        doc_number    = filing.get("Document Number", "")

        # ── Years in business ──
        years_in_business = ""
        if date_filed:
            try:
                filed_dt = datetime.strptime(date_filed, "%m/%d/%Y")
                years_in_business = str(
                    round((datetime.now() - filed_dt).days / 365.25, 1)
                )
            except Exception:
                pass

        def _section_text(title, html_body):
            """Extract visible text from a named detailSection."""
            m = re.search(
                rf"<span>\s*{re.escape(title)}\s*</span>(.*?)(?=<div[^>]*class=\"detailSection|$)",
                html_body, re.DOTALL | re.IGNORECASE,
            )
            if not m:
                return []
            chunk = m.group(1)
            chunk = re.sub(r"<br\s*/?>", "\n", chunk)
            chunk = re.sub(r"<[^>]+>", "", chunk)
            chunk = chunk.replace("&amp;", "&").replace("&nbsp;", " ")
            return [l.strip() for l in chunk.splitlines() if l.strip()]

        # ── Registered Agent ──
        ra_lines = _section_text("Registered Agent Name &amp; Address", html)
        reg_agent      = ra_lines[0] if ra_lines else ""
        reg_agent_addr = ", ".join(ra_lines[1:]) if len(ra_lines) > 1 else ""

        # ── Principal Address ──
        pa_lines       = _section_text("Principal Address", html)
        principal_addr = ", ".join(pa_lines)

        # ── Officers ──
        owner_name = ""
        off_m = re.search(
            r"<span>\s*Officer/Director Detail\s*</span>(.*?)(?=<div[^>]*class=\"detailSection|$)",
            html, re.DOTALL | re.IGNORECASE,
        )
        if off_m:
            off_chunk = re.sub(r"<br\s*/?>", "\n", off_m.group(1))
            off_chunk = re.sub(r"<[^>]+>", "\n", off_chunk)
            off_chunk = off_chunk.replace("&amp;", "&").replace("&nbsp;", " ")
            off_lines = [l.strip() for l in off_chunk.splitlines() if l.strip()]
            # Officer names come after a "Title X" line
            for i, line in enumerate(off_lines):
                if line.lower().startswith("title") and i + 1 < len(off_lines):
                    candidate = off_lines[i + 1]
                    # Must look like a name (all caps, letters)
                    if re.match(r"[A-Z][A-Z ,.\-']+$", candidate):
                        owner_name = candidate
                        break

        return {
            "found":             True,
            "sunbiz_url":        detail_url,
            "entity_type":       entity_type,
            "entity_name":       entity_name,
            "document_number":   doc_number,
            "date_filed":        date_filed,
            "years_in_business": years_in_business,
            "sunbiz_status":     status,
            "principal_address": principal_addr,
            "registered_agent":  reg_agent,
            "reg_agent_address": reg_agent_addr,
            "owner_name":        owner_name,
        }

    except Exception as e:
        return {"error": str(e), "searched": business_name}


def scrape_website_contact(url):
    """Visit a business website and extract email, Instagram, Facebook, phone."""
    if not url:
        return {"error": "No URL provided"}

    # Normalise URL
    if not url.startswith("http"):
        url = "https://" + url

    emails      = set()
    instagram   = ""
    facebook    = ""
    phones      = set()

    # Pages to attempt
    base = url.rstrip("/")
    pages = [base, base + "/contact", base + "/about", base + "/contact-us"]

    for page_url in pages:
        try:
            resp = requests.get(
                page_url, headers=SCRAPE_HEADERS,
                timeout=10, allow_redirects=True,
            )
            if resp.status_code >= 400:
                continue
            html = resp.text

            # ── Emails ──
            found = re.findall(
                r'\b[A-Za-z0-9._%+\-]+@[A-Za-z0-9.\-]+\.[A-Za-z]{2,}\b', html
            )
            noise = {
                "example", "sentry", "wixpress", "squarespace", "wordpress",
                "schema", "domain", "email", "support@sentry", "noreply",
                "webmaster", "user@",
            }
            for e in found:
                el = e.lower()
                if not any(n in el for n in noise) and len(el) < 80:
                    emails.add(el)

            # ── Instagram ──
            if not instagram:
                ig = re.search(
                    r'(?:href|content)="https?://(?:www\.)?instagram\.com/([^/"?#\s]+)',
                    html, re.IGNORECASE,
                )
                if ig and ig.group(1) not in ("p", "explore", "accounts", "stories"):
                    instagram = f"https://www.instagram.com/{ig.group(1)}"

            # ── Facebook ──
            if not facebook:
                fb = re.search(
                    r'(?:href|content)="https?://(?:www\.)?facebook\.com/([^/"?#\s]+)',
                    html, re.IGNORECASE,
                )
                if fb:
                    handle = fb.group(1)
                    skip = {"sharer", "share", "dialog", "plugins", "login", "groups", "events"}
                    if handle not in skip:
                        facebook = f"https://www.facebook.com/{handle}"

            # ── Phones ──
            tel_links = re.findall(r'href="tel:([^"]+)"', html, re.IGNORECASE)
            for p in tel_links:
                clean = re.sub(r"[^\d+]", "", p)
                if len(clean) >= 10:
                    phones.add(p.strip())

        except Exception:
            continue

    # If requests-based scraping found nothing useful, try Playwright for JS-heavy sites
    if not emails and not instagram and not facebook and not phones:
        try:
            from playwright.sync_api import sync_playwright
            import time as _time
            with sync_playwright() as pw:
                browser = pw.chromium.launch(headless=True, args=["--no-sandbox", "--disable-dev-shm-usage", "--disable-gpu"])
                ctx = browser.new_context(user_agent=SCRAPE_HEADERS["User-Agent"])
                pg = ctx.new_page()
                for page_url in pages[:2]:  # just home + /contact
                    try:
                        pg.goto(page_url, wait_until="domcontentloaded", timeout=15000)
                        _time.sleep(1)
                        html = pg.content()
                        found = re.findall(
                            r'\b[A-Za-z0-9._%+\-]+@[A-Za-z0-9.\-]+\.[A-Za-z]{2,}\b', html
                        )
                        noise = {
                            "example", "sentry", "wixpress", "squarespace", "wordpress",
                            "schema", "domain", "email", "support@sentry", "noreply",
                            "webmaster", "user@",
                        }
                        for e in found:
                            el = e.lower()
                            if not any(n in el for n in noise) and len(el) < 80:
                                emails.add(el)
                        if not instagram:
                            ig = re.search(
                                r'(?:href|content)="https?://(?:www\.)?instagram\.com/([^/"?#\s]+)',
                                html, re.IGNORECASE,
                            )
                            if ig and ig.group(1) not in ("p", "explore", "accounts", "stories"):
                                instagram = f"https://www.instagram.com/{ig.group(1)}"
                        if not facebook:
                            fb = re.search(
                                r'(?:href|content)="https?://(?:www\.)?facebook\.com/([^/"?#\s]+)',
                                html, re.IGNORECASE,
                            )
                            if fb:
                                handle = fb.group(1)
                                skip = {"sharer", "share", "dialog", "plugins", "login", "groups", "events"}
                                if handle not in skip:
                                    facebook = f"https://www.facebook.com/{handle}"
                        tel_links = re.findall(r'href="tel:([^"]+)"', html, re.IGNORECASE)
                        for p in tel_links:
                            clean = re.sub(r"[^\d+]", "", p)
                            if len(clean) >= 10:
                                phones.add(p.strip())
                        if emails or instagram or facebook or phones:
                            break
                    except Exception:
                        continue
                browser.close()
        except Exception:
            pass

    # Prefer info@, contact@, hello@ style emails as "general email"
    priority_prefixes = ("info", "contact", "hello", "office", "admin", "mail", "booking")
    general_email = ""
    for e in emails:
        if any(e.startswith(p + "@") for p in priority_prefixes):
            general_email = e
            break
    if not general_email and emails:
        general_email = sorted(emails)[0]

    return {
        "general_email":  general_email,
        "all_emails":     sorted(emails)[:6],
        "instagram_url":  instagram,
        "facebook_url":   facebook,
        "phones":         list(phones)[:3],
    }


def _hunter_domain_search(domain):
    """
    Silently look up emails for a domain using Hunter.io.
    Returns the best email found, or "" if nothing found or key not set.
    """
    global _hunter_key
    if not _hunter_key or not domain:
        return ""
    # Strip protocol/path — just need the bare domain
    domain = re.sub(r'^https?://', '', domain).split('/')[0].split('?')[0].strip()
    if not domain:
        return ""
    try:
        r = requests.get(
            "https://api.hunter.io/v2/domain-search",
            params={"domain": domain, "api_key": _hunter_key, "limit": 10},
            timeout=10,
        )
        data = r.json()
        emails = data.get("data", {}).get("emails", [])
        if not emails:
            return ""
        # Sort by confidence descending, prefer generic/owner type
        emails.sort(key=lambda e: (
            1 if e.get("type") in ("generic", "personal") else 0,
            e.get("confidence", 0)
        ), reverse=True)
        return emails[0].get("value", "")
    except Exception:
        return ""


def _google_places_lookup(business_name: str, city: str = "", state: str = "", address: str = ""):
    """
    Official Google Places API. Tries Places API (New) first — that's what
    Google enables for new projects. Falls back to the legacy Places API
    if the new one isn't enabled. Works from cloud hosts (Render, etc.)
    where headless Maps scraping is blocked.

    Set GOOGLE_PLACES_API_KEY or GOOGLE_MAPS_API_KEY (same key, either name).
    Enable in Google Cloud Console: "Places API (New)" — and/or "Places API".
    """
    import logging as _logging
    log = _logging.getLogger(__name__)

    key = (os.environ.get("GOOGLE_PLACES_API_KEY") or os.environ.get("GOOGLE_MAPS_API_KEY") or "").strip()
    if not key or not (business_name or "").strip():
        return {}
    parts = [(business_name or "").strip()]
    if (address or "").strip():
        parts.append(address.strip())
    else:
        if city:  parts.append(city)
        if state: parts.append(state)
    query = ", ".join(p for p in parts if p) if address else " ".join(p for p in parts if p)
    if not query:
        return {}

    out = {
        "google_rating":       "",
        "google_review_count": "",
        "website":             "",
        "address":             "",
        "business_phone":      "",
    }

    # ── Strategy 1: Places API (New) — preferred ──────────────────────────
    new_api_unavailable = False
    try:
        r = requests.post(
            "https://places.googleapis.com/v1/places:searchText",
            json={"textQuery": query, "maxResultCount": 5},
            headers={
                "Content-Type":     "application/json",
                "X-Goog-Api-Key":   key,
                "X-Goog-FieldMask": (
                    "places.id,places.displayName,places.formattedAddress,places.rating,"
                    "places.userRatingCount,places.websiteUri,places.nationalPhoneNumber,"
                    "places.internationalPhoneNumber"
                ),
            },
            timeout=20,
        )
        if r.status_code == 200:
            data = r.json() or {}
            places = data.get("places") or []
            if places:
                best = _pick_best_place_match(places, business_name)
                if best.get("rating") is not None:
                    out["google_rating"] = str(best["rating"])
                if best.get("userRatingCount") is not None:
                    out["google_review_count"] = str(best["userRatingCount"])
                w = (best.get("websiteUri") or "").strip()
                if w:
                    out["website"] = w.split("?")[0].rstrip("/")
                addr = (best.get("formattedAddress") or "").strip()
                if addr:
                    out["address"] = addr
                ph = (best.get("nationalPhoneNumber") or best.get("internationalPhoneNumber") or "").strip()
                if ph:
                    out["business_phone"] = ph
                return out
            return out  # OK but zero results
        else:
            try:
                err = (r.json() or {}).get("error", {})
            except Exception:
                err = {}
            err_status = err.get("status", "")
            err_msg    = err.get("message", "")
            log.warning("[places-new] HTTP %s status=%s msg=%s", r.status_code, err_status, err_msg[:200])
            # Detect "API not enabled" so we can fall back to the legacy endpoint cleanly
            if r.status_code in (403, 404) or err_status in ("PERMISSION_DENIED", "FAILED_PRECONDITION"):
                new_api_unavailable = True
            else:
                return {"_error": f"Places API (New) HTTP {r.status_code}: {err_status} {err_msg}", **out}
    except Exception as e:
        log.warning("[places-new] exception: %s", e)
        new_api_unavailable = True  # try legacy

    # ── Strategy 2: Legacy Places API (Text Search + Details) ─────────────
    if new_api_unavailable:
        try:
            r = requests.get(
                "https://maps.googleapis.com/maps/api/place/textsearch/json",
                params={"query": query, "key": key},
                timeout=20,
            )
            data = r.json() if r.status_code == 200 else {}
            status = data.get("status", f"HTTP {r.status_code}")
            if status not in ("OK", "ZERO_RESULTS"):
                msg = data.get("error_message", "")
                return {"_error": f"Places (legacy) Text Search: {status} {msg}".strip(), **out}
            results = data.get("results") or []
            if not results:
                return out
            best = results[0]
            place_id = best.get("place_id")
            if best.get("rating") is not None:
                out["google_rating"] = str(best["rating"])
            if best.get("user_ratings_total") is not None:
                out["google_review_count"] = str(best["user_ratings_total"])
            out["address"] = (best.get("formatted_address") or "").strip()
            if place_id:
                r2 = requests.get(
                    "https://maps.googleapis.com/maps/api/place/details/json",
                    params={
                        "place_id": place_id,
                        "fields": "website,formatted_phone_number,international_phone_number,rating,user_ratings_total",
                        "key":    key,
                    },
                    timeout=20,
                )
                d2 = r2.json() if r2.status_code == 200 else {}
                if d2.get("status") == "OK":
                    res = d2.get("result") or {}
                    if res.get("rating") is not None:
                        out["google_rating"] = str(res["rating"])
                    if res.get("user_ratings_total") is not None:
                        out["google_review_count"] = str(res["user_ratings_total"])
                    w = (res.get("website") or "").strip()
                    if w:
                        out["website"] = w.split("?")[0].rstrip("/")
                    ph = (res.get("formatted_phone_number") or res.get("international_phone_number") or "").strip()
                    if ph:
                        out["business_phone"] = ph
        except Exception as e:
            return {"_error": str(e), **out}

    return out


def _pick_best_place_match(places: list, business_name: str) -> dict:
    """Pick the place whose displayName best matches the queried business name."""
    if not places:
        return {}
    if len(places) == 1:
        return places[0]
    target = (business_name or "").lower()
    target_words = set(re.sub(r"[^a-z0-9\s]", "", target).split())
    target_words -= {"the", "a", "an", "of", "and", "&", "llc", "inc", "corp", "co"}
    best = places[0]
    best_score = -1.0
    for p in places:
        cand = (p.get("displayName", {}).get("text") or "").lower()
        cand_words = set(re.sub(r"[^a-z0-9\s]", "", cand).split())
        score = len(target_words & cand_words)
        # Bump score if rating exists (means it's a real, mappable business)
        if p.get("rating") is not None:
            score += 0.1
        if score > best_score:
            best_score = score
            best = p
    return best


def _outbound_proxy_url() -> str:
    """Optional outbound proxy URL for Playwright + requests.
    Set OUTBOUND_PROXY_URL (e.g. http://user:pass@proxy.example.com:8080) on hosts where
    Google blocks direct datacenter IPs (Render, Fly, etc.). Returns "" if not configured."""
    return (os.environ.get("OUTBOUND_PROXY_URL") or "").strip()


def _proxy_dict_for_requests():
    """`requests`-style proxy dict, or None."""
    p = _outbound_proxy_url()
    if not p:
        return None
    return {"http": p, "https": p}


def _proxy_for_playwright():
    """Playwright proxy kwarg, or None. Parses optional user:pass@host:port form."""
    p = _outbound_proxy_url()
    if not p:
        return None
    from urllib.parse import urlparse
    try:
        u = urlparse(p)
        cfg = {"server": f"{u.scheme}://{u.hostname}:{u.port}" if u.port else f"{u.scheme}://{u.hostname}"}
        if u.username:
            cfg["username"] = u.username
        if u.password:
            cfg["password"] = u.password
        return cfg
    except Exception:
        return {"server": p}


def _maps_search_url(query: str) -> str:
    # hl=en&gl=us forces English UI so our selectors and aria-labels match.
    return f"https://www.google.com/maps/search/{quote_plus(query)}?hl=en&gl=us"


def _dismiss_google_consent(page):
    """Click through Google's 'Before you continue' / cookie consent page if present.
    Without this, headless Chromium often gets stuck on the consent wall and
    the Maps panel never loads, so every extractor returns blank."""
    try:
        for sel in (
            'button[aria-label*="Accept all" i]',
            'button[aria-label*="Agree" i]',
            'button:has-text("Accept all")',
            'button:has-text("I agree")',
            'form[action*="consent"] button',
            'button[jsname="b3VHJd"]',
        ):
            btn = page.query_selector(sel)
            if btn:
                try:
                    btn.click(timeout=2000)
                    page.wait_for_load_state("domcontentloaded", timeout=8000)
                    return True
                except Exception:
                    continue
    except Exception:
        pass
    return False


def get_google_reviews(business_name, city="", state="", address_hint=""):
    """
    Rating, review count, website, address, phone from Google.

    1) If GOOGLE_PLACES_API_KEY (or GOOGLE_MAPS_API_KEY) is set, use the official
       Places API first — reliable on Render/cloud where headless Maps is often blocked.
    2) Otherwise (or to fill gaps), use Playwright + maps.google.com as before.
    """
    from playwright.sync_api import sync_playwright
    import logging as _logging
    log = _logging.getLogger(__name__)

    rating  = ""
    count   = ""
    website = ""
    address = ""
    phone   = ""
    html    = ""

    # Official API path (recommended for production / Render)
    places_key = (os.environ.get("GOOGLE_PLACES_API_KEY") or os.environ.get("GOOGLE_MAPS_API_KEY") or "").strip()
    if places_key:
        pl = _google_places_lookup(business_name, city, state, address=address_hint or "")
        if pl.get("_error"):
            log.warning("[maps] Places API: %s", pl["_error"])
        for k in ("google_rating", "google_review_count", "website", "address", "business_phone"):
            if pl.get(k):
                if k == "google_rating":
                    rating = pl[k]
                elif k == "google_review_count":
                    count = pl[k]
                elif k == "website":
                    website = pl[k]
                elif k == "address":
                    address = pl[k]
                elif k == "business_phone":
                    phone = pl[k]
        # If we got rating + website from Places, skip headless Maps (often blocked on cloud IPs)
        if rating and website:
            return {
                "google_rating":       rating,
                "google_review_count": count,
                "website":             website,
                "address":             address,
                "business_phone":      phone,
            }
        # Partial data: still try Playwright below for missing fields

    queries = []
    base = (business_name or "").strip()
    loc  = " ".join([p for p in [city, state] if p]).strip()
    if base and loc:
        queries.append(f"{base} {loc}")
    if base and city:
        queries.append(f"{base} {city}")
    if base and state and not loc:
        queries.append(f"{base} {state}")
    if base:
        queries.append(base)
    seen_q = set()
    queries = [q for q in queries if not (q in seen_q or seen_q.add(q))]

    def _scrape(query):
        nonlocal rating, count, website, address, phone, html
        try:
            with sync_playwright() as pw:
                launch_kwargs = dict(
                    headless=True,
                    args=[
                        "--no-sandbox",
                        "--disable-dev-shm-usage",
                        "--disable-gpu",
                        "--disable-blink-features=AutomationControlled",
                    ],
                )
                proxy_cfg = _proxy_for_playwright()
                if proxy_cfg:
                    launch_kwargs["proxy"] = proxy_cfg
                browser = pw.chromium.launch(**launch_kwargs)
                context = browser.new_context(
                    user_agent=(
                        "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) "
                        "AppleWebKit/537.36 (KHTML, like Gecko) "
                        "Chrome/120.0.0.0 Safari/537.36"
                    ),
                    viewport={"width": 1366, "height": 900},
                    locale="en-US",
                )
                # Pre-set the CONSENT cookie so Google skips the consent wall.
                try:
                    context.add_cookies([
                        {
                            "name":   "CONSENT",
                            "value":  "YES+cb.20210720-07-p0.en+FX+410",
                            "domain": ".google.com",
                            "path":   "/",
                        },
                        {
                            "name":   "SOCS",
                            "value":  "CAESEwgDEgk0ODE3Nzg3MjQaAmVuIAEaBgiA_LyaBg",
                            "domain": ".google.com",
                            "path":   "/",
                        },
                    ])
                except Exception:
                    pass

                page = context.new_page()
                page.goto(
                    _maps_search_url(query),
                    wait_until="domcontentloaded", timeout=30000,
                )
                page.wait_for_timeout(2000)

                # Handle consent wall if cookies didn't suppress it.
                if "consent.google.com" in (page.url or "") or page.query_selector('form[action*="consent"]'):
                    if _dismiss_google_consent(page):
                        page.wait_for_timeout(2000)
                    # Re-navigate to Maps after consent.
                    try:
                        page.goto(_maps_search_url(query), wait_until="domcontentloaded", timeout=30000)
                        page.wait_for_timeout(2000)
                    except Exception:
                        pass

                first = page.query_selector("a.hfpxzc")
                if first:
                    try:
                        first.click()
                    except Exception:
                        pass
                    page.wait_for_timeout(4500)

                html = page.content()

                if not website:
                    try:
                        for sel in (
                            'a[data-item-id="authority"]',
                            'a[aria-label*="website" i]',
                            'a[data-tooltip*="website" i]',
                        ):
                            el = page.query_selector(sel)
                            if el:
                                href = el.get_attribute("href") or ""
                                if href and not href.startswith("https://www.google"):
                                    website = href.split("?")[0].rstrip("/")
                                    break
                    except Exception:
                        pass

                if not address:
                    try:
                        for sel in (
                            '[data-item-id="address"] .Io6YTe',
                            'button[data-item-id="address"] .Io6YTe',
                        ):
                            el = page.query_selector(sel)
                            if el:
                                address = (el.inner_text() or "").strip()
                                if address:
                                    break
                    except Exception:
                        pass

                if not phone:
                    try:
                        for sel in (
                            '[data-item-id*="phone:tel"] .Io6YTe',
                            'button[data-tooltip="Copy phone number"] .Io6YTe',
                        ):
                            el = page.query_selector(sel)
                            if el:
                                phone = (el.inner_text() or "").strip()
                                if phone:
                                    break
                    except Exception:
                        pass

                for el in page.query_selector_all('[aria-label]'):
                    try:
                        lbl = el.get_attribute("aria-label") or ""
                        if not rating:
                            rm = re.search(r'(\d[\.,]\d)\s*stars?', lbl, re.IGNORECASE)
                            if rm:
                                rating = rm.group(1).replace(",", ".")
                            else:
                                rm2 = re.search(r'^(\d)\s*stars?$', lbl.strip(), re.IGNORECASE)
                                if rm2:
                                    rating = rm2.group(1)
                        if not count:
                            cm = re.search(r'([\d,]+)\s*reviews?', lbl, re.IGNORECASE)
                            if cm:
                                count = cm.group(1).replace(",", "")
                        if rating and count:
                            break
                    except Exception:
                        continue

                # Reviews: explicitly look for the parenthesized count next to the rating,
                # e.g. "4.5 stars (1,234)". Maps often shows it that way without "reviews".
                if not count:
                    try:
                        for sel in ('div.F7nice', 'span.F7nice', 'div.fontBodyMedium'):
                            el = page.query_selector(sel)
                            if el:
                                txt = (el.inner_text() or "").strip()
                                cm = re.search(r'\(([\d,]+)\)', txt)
                                if cm:
                                    count = cm.group(1).replace(",", "")
                                    break
                    except Exception:
                        pass

                if not rating:
                    for sel in ('span.ceNzKf', 'div.F7nice > span', 'span.fontBodyMedium'):
                        try:
                            el = page.query_selector(sel)
                            if el:
                                txt = (el.inner_text() or "").strip()
                                rm = re.search(r'(\d[\.,]\d)', txt)
                                if rm:
                                    rating = rm.group(1).replace(",", ".")
                                    break
                        except Exception:
                            continue

                browser.close()
        except Exception as e:
            log.warning("[maps] scrape failed for %r: %s", query, e)

        if not rating:
            rm = re.search(r'(\d[\.,]\d)\s*stars?', html or "", re.IGNORECASE)
            if rm:
                rating = rm.group(1).replace(",", ".")
        if not count:
            cm = re.search(r'([\d,]+)\s*reviews?', html or "", re.IGNORECASE)
            if cm:
                count = cm.group(1).replace(",", "")
            else:
                # Look for the standalone "(1,234)" pattern that appears near the rating
                if rating:
                    rating_pos = (html or "").find(rating)
                    if rating_pos > 0:
                        nearby = (html or "")[rating_pos:rating_pos + 600]
                        cm2 = re.search(r'\(([\d,]+)\)', nearby)
                        if cm2:
                            count = cm2.group(1).replace(",", "")

    for q in queries:
        _scrape(q)
        if rating and count and website:
            break

    # If Maps failed to find a website, try a lightweight web-search fallback.
    if not website and base:
        try:
            website = _discover_website_via_search(base, city, state)
        except Exception:
            pass

    return {
        "google_rating":       rating,
        "google_review_count": count,
        "website":             website,
        "address":             address,
        "business_phone":      phone,
    }


_BAD_WEBSITE_DOMAINS = {
    "google.com", "google.co", "facebook.com", "instagram.com", "twitter.com",
    "x.com", "tiktok.com", "linkedin.com", "youtube.com", "yelp.com",
    "tripadvisor.com", "wikipedia.org", "duckduckgo.com", "bing.com", "maps.app.goo.gl",
    "mapquest.com", "yellowpages.com", "bbb.org", "indeed.com", "glassdoor.com",
    "opentable.com", "doordash.com", "ubereats.com", "grubhub.com", "groupon.com",
    "thumbtack.com", "angi.com", "houzz.com", "pinterest.com", "reddit.com",
    "apple.com", "amazon.com", "ebay.com", "etsy.com",
}


def _filter_candidate_url(target: str) -> str:
    """Return the URL if it points to a real business site, else empty string."""
    from urllib.parse import urlparse as _urlparse
    try:
        host = (_urlparse(target).netloc or "").lower()
        if host.startswith("www."):
            host = host[4:]
        if not host or "." not in host:
            return ""
        # Strip subdomain so we match e.g. "search.yelp.com" against "yelp.com"
        host_parts = host.split(".")
        base_host = ".".join(host_parts[-2:])
        if base_host in _BAD_WEBSITE_DOMAINS:
            return ""
        # Exclude obvious social/profile paths
        path = (_urlparse(target).path or "").lower()
        if any(p in path for p in ("/maps/", "/profile", "/pages/", "/business/")):
            return ""
        return target.split("?")[0].rstrip("/")
    except Exception:
        return ""


def _discover_website_via_requests(query: str) -> str:
    """Quick, no-browser website discovery via Brave Search and DuckDuckGo HTML.
    Returns "" if blocked. Used as a fast first attempt before Playwright."""
    headers = {
        "User-Agent": (
            "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) "
            "AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36"
        ),
        "Accept-Language": "en-US,en;q=0.9",
        "Accept": "text/html,application/xhtml+xml,*/*;q=0.8",
    }
    from urllib.parse import unquote as _unquote
    proxies = _proxy_dict_for_requests()

    # 1) Brave Search HTML — usually returns real results to scrapers
    try:
        r = requests.get(
            "https://search.brave.com/search",
            params={"q": query, "source": "web"},
            headers=headers, timeout=10, proxies=proxies,
        )
        if r.status_code == 200:
            for raw in re.findall(r'<a [^>]*href="(https?://[^"]+)"[^>]*class="[^"]*result-header', r.text):
                site = _filter_candidate_url(raw)
                if site:
                    return site
            for raw in re.findall(r'<a [^>]*class="[^"]*result-header[^"]*"[^>]*href="(https?://[^"]+)"', r.text):
                site = _filter_candidate_url(raw)
                if site:
                    return site
    except Exception:
        pass

    # 2) DuckDuckGo lite (often works when html.duckduckgo doesn't)
    try:
        r = requests.get(
            "https://lite.duckduckgo.com/lite/",
            params={"q": query},
            headers=headers, timeout=10, proxies=proxies,
        )
        if r.status_code == 200:
            for raw in re.findall(r'<a [^>]*href="(https?://[^"]+)"', r.text):
                if "duckduckgo.com" in raw:
                    continue
                site = _filter_candidate_url(raw)
                if site:
                    return site
    except Exception:
        pass

    return ""


def _discover_website_via_search(business_name: str, city: str = "", state: str = "") -> str:
    """Best-effort website discovery. Tries fast `requests`-based search first
    (Brave / DDG lite), then falls back to a Playwright-driven Google search.
    Returns a normalized URL or empty string; skips directory/social domains."""
    if not business_name:
        return ""

    parts = [business_name]
    if city:  parts.append(city)
    if state: parts.append(state)
    query = " ".join(parts) + " official site"

    # Fast path: no headless browser
    fast = _discover_website_via_requests(query)
    if fast:
        return fast

    # Slow path: full Google search via Playwright
    from playwright.sync_api import sync_playwright

    try:
        with sync_playwright() as pw:
            _slow_proxy = _proxy_for_playwright()
            _slow_kwargs = dict(
                headless=True,
                args=[
                    "--no-sandbox",
                    "--disable-dev-shm-usage",
                    "--disable-gpu",
                    "--disable-blink-features=AutomationControlled",
                ],
            )
            if _slow_proxy:
                _slow_kwargs["proxy"] = _slow_proxy
            browser = pw.chromium.launch(**_slow_kwargs)
            context = browser.new_context(
                user_agent=(
                    "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) "
                    "AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36"
                ),
                viewport={"width": 1366, "height": 900},
                locale="en-US",
            )
            try:
                context.add_cookies([
                    {"name": "CONSENT", "value": "YES+cb.20210720-07-p0.en+FX+410", "domain": ".google.com", "path": "/"},
                    {"name": "SOCS", "value": "CAESEwgDEgk0ODE3Nzg3MjQaAmVuIAEaBgiA_LyaBg", "domain": ".google.com", "path": "/"},
                ])
            except Exception:
                pass

            page = context.new_page()
            page.goto(
                f"https://www.google.com/search?q={quote_plus(query)}&hl=en&gl=us",
                wait_until="domcontentloaded", timeout=25000,
            )
            page.wait_for_timeout(1500)

            if "consent.google.com" in (page.url or "") or page.query_selector('form[action*="consent"]'):
                _dismiss_google_consent(page)
                try:
                    page.goto(
                        f"https://www.google.com/search?q={quote_plus(query)}&hl=en&gl=us",
                        wait_until="domcontentloaded", timeout=25000,
                    )
                    page.wait_for_timeout(1500)
                except Exception:
                    pass

            html_text = page.content()
            browser.close()

        # Google result links use class g and h3 inside an <a href="..."> wrapper
        # Extract every external href and pick the first non-directory match.
        from urllib.parse import urlparse as _urlparse, parse_qs as _parse_qs

        candidates = []
        for raw in re.findall(r'href="(/url\?[^"]+)"', html_text):
            try:
                qs = _parse_qs(_urlparse(raw).query)
                if "q" in qs and qs["q"][0].startswith("http"):
                    candidates.append(qs["q"][0])
            except Exception:
                continue
        for raw in re.findall(r'href="(https?://[^"]+)"', html_text):
            candidates.append(raw)

        for raw in candidates:
            site = _filter_candidate_url(raw)
            if site:
                return site
    except Exception:
        pass

    return ""


def hubspot_create_contact(email, first_name="", last_name="", company="",
                            phone="", website="", job_title="", linkedin="",
                            _hubspot_token=None):
    hubspot_token = _hubspot_token or os.getenv("HUBSPOT_TOKEN", "")
    if not hubspot_token:
        return {"error": "HubSpot token not configured. Please set it in Settings."}
    properties = {"email": email}
    if first_name: properties["firstname"]    = first_name
    if last_name:  properties["lastname"]     = last_name
    if company:    properties["company"]      = company
    if phone:      properties["phone"]        = phone
    if website:    properties["website"]      = website
    if job_title:  properties["jobtitle"]     = job_title
    if linkedin:   properties["linkedin_bio"] = linkedin
    try:
        r = requests.post(
            "https://api.hubapi.com/crm/v3/objects/contacts",
            json={"properties": properties},
            headers={
                "Authorization": f"Bearer {hubspot_token.strip()}",
                "Content-Type":  "application/json",
            },
            timeout=20,
        )
        data = r.json()
        if r.status_code in (200, 201):
            return {"success": True, "id": data.get("id"), "email": email, "company": company}
        elif r.status_code == 409:
            return {"success": False, "error": "Contact already exists", "email": email}
        else:
            return {"success": False, "error": data.get("message", str(data)), "email": email}
    except Exception as e:
        return {"error": str(e)}


def get_collected_leads():
    """Return leads already collected in this session."""
    global _leads_store
    if not _leads_store:
        return {"leads": [], "count": 0, "message": "No leads collected yet in this session."}
    return {"leads": _leads_store, "count": len(_leads_store)}


def upload_leads_to_hubspot(_hubspot_token=None):
    """
    Upload all collected leads to HubSpot CRM in one call.
    Handles field mapping automatically: splits owner name, picks best email,
    uses trade_name as company. Returns a summary of uploaded contacts.
    """
    global _leads_store
    hubspot_token = _hubspot_token or os.getenv("HUBSPOT_TOKEN", "")
    if not hubspot_token:
        return {"error": "HubSpot token not configured. Please set it in Settings."}
    if not _leads_store:
        return {"error": "No leads collected yet. Run a lead search first."}

    # Only process leads that have at least one email address
    leads_with_email = [
        l for l in _leads_store
        if l.get("owner_email") or l.get("general_email") or l.get("reg_agent_email")
    ]
    skipped_no_email = len(_leads_store) - len(leads_with_email)
    requested_upload = len(leads_with_email)
    leads_with_email = leads_with_email[:HUBSPOT_PUSH_LIMIT]
    skipped_over_limit = max(0, requested_upload - HUBSPOT_PUSH_LIMIT)

    results = {
        "uploaded": 0, "skipped": skipped_no_email + skipped_over_limit, "errors": [],
        "contacts": [],
        "no_email_count": skipped_no_email,
        "requested": requested_upload,
        "limit_applied": HUBSPOT_PUSH_LIMIT,
        "capped": requested_upload > HUBSPOT_PUSH_LIMIT,
        "skipped_over_limit": skipped_over_limit,
    }

    for lead in leads_with_email:
        # ── Pick best email: owner → general → registered agent → placeholder ──
        email = (
            lead.get("owner_email") or
            lead.get("general_email") or
            lead.get("reg_agent_email") or
            "johndoe@gmail.com"
        ).strip()

        # ── Split owner name into first / last (fall back to registered agent) ──
        raw_name = (lead.get("owner_name") or lead.get("registered_agent") or "").strip()
        # Guard: skip if the "name" is actually an email address
        if "@" in raw_name or re.match(r'^[\w._%+\-]+@[\w.\-]+\.[a-z]{2,}$', raw_name, re.IGNORECASE):
            raw_name = ""
        # Names from Sunbiz are "LAST, FIRST MIDDLE" or "FIRST LAST"
        first_name, last_name = "", ""
        if raw_name:
            if "," in raw_name:
                parts = [p.strip().title() for p in raw_name.split(",", 1)]
                last_name  = parts[0]
                first_name = parts[1].split()[0] if parts[1] else ""
            else:
                parts = raw_name.title().split()
                first_name = parts[0] if parts else ""
                last_name  = " ".join(parts[1:]) if len(parts) > 1 else ""
        # Final guard: ensure neither first nor last name contains an @
        if "@" in first_name: first_name = ""
        if "@" in last_name:  last_name  = ""

        # ── Company name — prefer trade name (human-readable) over entity name ──
        company = (lead.get("trade_name") or lead.get("entity_name") or "").strip()

        # ── Phone — prefer owner cell, fall back to business phone ──
        phone = (lead.get("owner_phone") or lead.get("business_phone") or "").strip()

        properties = {"email": email}
        if first_name: properties["firstname"]  = first_name
        if last_name:  properties["lastname"]   = last_name
        if company:    properties["company"]    = company
        if phone:      properties["phone"]      = phone
        if lead.get("website"):  properties["website"]  = lead["website"]
        properties["jobtitle"] = "Owner"

        try:
            r = requests.post(
                "https://api.hubapi.com/crm/v3/objects/contacts",
                json={"properties": properties},
                headers={
                    "Authorization": f"Bearer {hubspot_token.strip()}",
                    "Content-Type":  "application/json",
                },
                timeout=20,
            )
            data = r.json()
            if r.status_code in (200, 201):
                results["uploaded"] += 1
                results["contacts"].append({"email": email, "company": company, "id": data.get("id")})
            elif r.status_code == 409:
                results["skipped"] += 1
                results["errors"].append(f"Already exists: {email}")
            else:
                results["skipped"] += 1
                results["errors"].append(f"{email}: {data.get('message', str(r.status_code))}")
        except Exception as e:
            results["skipped"] += 1
            results["errors"].append(f"{email}: {str(e)}")

    return results


def save_leads_csv(leads):
    """Save enriched leads list to leads.csv."""
    global _leads_store
    _leads_store = leads
    _save_leads_to_file(leads)
    return {"success": True, "count": len(leads), "path": "leads.csv"}


def _save_leads_to_file(leads):
    try:
        path = os.path.join(os.path.dirname(__file__), "leads.csv")
        with open(path, "w", newline="") as f:
            writer = csv.DictWriter(f, fieldnames=LEAD_FIELDS, extrasaction="ignore")
            writer.writeheader()
            writer.writerows(leads)
    except Exception:
        pass


_research_report_store = {"markdown": "", "title": "", "ts": ""}


def save_research_report(report_markdown: str, title: str = ""):
    """Save the in-depth research report markdown so the user can download it.
    The model writes the markdown; this tool just persists it and returns a link."""
    global _research_report_store
    md = (report_markdown or "").strip()
    if not md:
        return {"error": "Empty report. Pass the full markdown report as `report_markdown`."}
    _research_report_store = {
        "markdown": md,
        "title":    (title or "Tenant Prospecting Research Report").strip(),
        "ts":       datetime.now().isoformat(timespec="seconds"),
    }
    try:
        path = os.path.join(os.path.dirname(__file__), "research_report.md")
        with open(path, "w", encoding="utf-8") as f:
            f.write(md)
    except Exception:
        pass
    return {
        "success":      True,
        "title":        _research_report_store["title"],
        "char_count":   len(md),
        "download_url": "/api/download/report",
    }


def save_outreach_csv(drafts):
    global _outreach_store
    _outreach_store = drafts
    try:
        path = os.path.join(os.path.dirname(__file__), "outreach_drafts.csv")
        with open(path, "w", newline="") as f:
            writer = csv.DictWriter(
                f, fieldnames=["name", "email", "subject_line", "email_body"],
                extrasaction="ignore",
            )
            writer.writeheader()
            writer.writerows(drafts)
        return {"success": True, "path": path, "count": len(drafts), "drafts": drafts}
    except Exception as e:
        return {"error": str(e)}


def _get_gmail_creds():
    """Get Gmail address and app password — checks session, .env, then persistent file."""
    import flask
    _gmail_app_file = os.path.join(os.path.dirname(os.path.abspath(__file__)), ".gmail_app.json")

    # 1. Try session
    try:
        gmail_address  = flask.session.get("gmail_address", "")
        gmail_password = flask.session.get("gmail_app_password", "")
    except RuntimeError:
        gmail_address, gmail_password = "", ""

    # 2. Fall back to .env
    if not gmail_address:
        gmail_address  = os.getenv("GMAIL_ADDRESS", "")
    if not gmail_password:
        gmail_password = os.getenv("GMAIL_APP_PASSWORD", "")

    # 3. Fall back to persistent file (saved via Settings UI)
    if (not gmail_address or not gmail_password) and os.path.exists(_gmail_app_file):
        try:
            with open(_gmail_app_file) as f:
                data = json.load(f)
            gmail_address  = gmail_address  or data.get("gmail_address", "")
            gmail_password = gmail_password or data.get("gmail_app_password", "")
        except Exception:
            pass

    return gmail_address, gmail_password


def send_gmail_email(to=None, subject=None, body=None, send_all_drafts=False):
    """Send email(s) via Gmail SMTP using an App Password."""
    import smtplib
    gmail_address, gmail_password = _get_gmail_creds()
    if not gmail_address or not gmail_password:
        return {"error": "Gmail not configured. Add your Gmail address and App Password in Settings."}

    def _send_one(to_addr, subj, text_body):
        msg = email_lib.mime.text.MIMEText(text_body, "plain")
        msg["From"]    = gmail_address
        msg["To"]      = to_addr
        msg["Subject"] = subj
        with smtplib.SMTP_SSL("smtp.gmail.com", 465) as server:
            server.login(gmail_address, gmail_password)
            server.send_message(msg)

    if send_all_drafts:
        global _outreach_store
        if not _outreach_store:
            return {"error": "No outreach drafts found. Write outreach emails first."}
        sent, failed = [], []
        for draft in _outreach_store:
            try:
                _send_one(draft.get("email", ""), draft.get("subject_line", "No subject"), draft.get("email_body", ""))
                sent.append(draft.get("email", ""))
            except Exception as ex:
                failed.append({"email": draft.get("email", ""), "error": str(ex)})
        return {"success": True, "sent_count": len(sent), "failed_count": len(failed), "sent": sent, "failed": failed}
    else:
        if not to or not subject or not body:
            return {"error": "Missing required fields: to, subject, body"}
        try:
            _send_one(to, subject, body)
            return {"success": True, "to": to}
        except Exception as e:
            return {"error": f"Gmail send failed: {str(e)}"}


def create_gmail_drafts(draft_all=True, to=None, subject=None, body=None):
    """Save outreach emails as drafts in Gmail using the Gmail API (requires OAuth) or IMAP APPEND."""
    # For App Password users, we use the Gmail API with basic auth via IMAP to append to Drafts
    import smtplib, imaplib
    gmail_address, gmail_password = _get_gmail_creds()
    if not gmail_address or not gmail_password:
        return {"error": "Gmail not configured. Add your Gmail address and App Password in Settings."}

    def _append_draft(to_addr, subj, text_body):
        msg = email_lib.mime.text.MIMEText(text_body, "plain")
        msg["From"]    = gmail_address
        msg["To"]      = to_addr
        msg["Subject"] = subj
        raw = msg.as_bytes()
        imap = imaplib.IMAP4_SSL("imap.gmail.com")
        imap.login(gmail_address, gmail_password)
        imap.append("[Gmail]/Drafts", "\\Draft", imaplib.Time2Internaldate(datetime.now().timestamp()), raw)
        imap.logout()

    if draft_all:
        global _outreach_store
        if not _outreach_store:
            return {"error": "No outreach drafts found. Write outreach emails first, then create Gmail drafts."}
        created, failed = [], []
        for draft in _outreach_store:
            try:
                _append_draft(draft.get("email", ""), draft.get("subject_line", "No subject"), draft.get("email_body", ""))
                created.append(draft.get("email", ""))
            except Exception as ex:
                failed.append({"email": draft.get("email", ""), "error": str(ex)})
        return {
            "success": True,
            "created_count": len(created),
            "failed_count": len(failed),
            "message": f"Saved {len(created)} email(s) to your Gmail Drafts folder. Open Gmail to review and send.",
        }
    else:
        if not to or not subject or not body:
            return {"error": "Missing required fields: to, subject, body"}
        try:
            _append_draft(to, subject, body)
            return {"success": True, "to": to, "message": "Draft saved to your Gmail Drafts folder."}
        except Exception as e:
            return {"error": f"Gmail draft creation failed: {str(e)}"}


def _find_person_contact(name, business_name="", city="", state=""):
    """
    Web-search for a person's email and phone number.
    Used for owner and registered agent contact lookup.
    Returns {"email": "...", "phone": "..."}.
    """
    if not name:
        return {"email": "", "phone": ""}

    noise_domains = {"example", "sentry", "wixpress", "squarespace", "domain", "noreply",
                     "wordpress", "schema", "w3.org", "google", "yelp", "facebook"}

    def _extract(text):
        emails = re.findall(r'\b[A-Za-z0-9._%+\-]+@[A-Za-z0-9.\-]+\.[A-Za-z]{2,}\b', text)
        email = ""
        for e in emails:
            el = e.lower()
            if not any(n in el for n in noise_domains) and len(el) < 60:
                email = el
                break
        phones = re.findall(r'\(?\d{3}\)?[\s\.\-]?\d{3}[\s\.\-]?\d{4}', text)
        phone = phones[0].strip() if phones else ""
        return email, phone

    # Try multiple query strategies
    queries = []
    if business_name:
        queries.append(f'"{name}" "{business_name}" email contact')
        queries.append(f'"{name}" {city} {state} {business_name} owner email')
    queries.append(f'"{name}" {city} {state} email phone')

    for query in queries:
        try:
            result = web_search(query.strip())
            text = result.get("results", "")
            email, phone = _extract(text)
            if email or phone:
                return {"email": email, "phone": phone}
        except Exception:
            continue

    return {"email": "", "phone": ""}


def enrich_leads_batch(leads=None):
    """
    Enrich every lead in the list with Sunbiz, website contact info, and
    Google Maps reviews — all in one tool call.  Yields progress via a
    shared list; returns the fully enriched leads list and saves to CSV.
    """
    import concurrent.futures
    import logging as _logging
    _elog = _logging.getLogger(__name__)
    global _leads_store

    # Be tolerant if the model forgets to pass leads:
    # fall back to the currently collected leads in session memory.
    if leads is None:
        leads = _leads_store
    if not leads:
        return {
            "error": (
                "No leads provided to enrich. Please run a lead search first "
                "or pass a `leads` array into enrich_leads_batch."
            )
        }

    enriched = []

    def _enrich_one(lead):
        result = dict(lead)
        for f in LEAD_FIELDS:
            result.setdefault(f, "")

        name  = result.get("trade_name", "")
        city  = result.get("city", "")
        state = result.get("state", "")

        sources = {
            "sunbiz":  False,
            "google":  False,
            "website": False,
            "apollo":  False,
        }

        # 1. Google Maps FIRST — gives us website, address, phone, rating, reviews.
        #    Doing this first means subsequent steps (website scrape, Apollo) work
        #    even when the lead came in without a website URL.
        if not (result.get("google_rating") and result.get("google_review_count") and result.get("website")):
            try:
                gr = get_google_reviews(name, city, state, address_hint=result.get("address") or "")
                if gr.get("google_rating"):
                    result["google_rating"] = gr["google_rating"]
                    sources["google"] = True
                if gr.get("google_review_count"):
                    result["google_review_count"] = gr["google_review_count"]
                    sources["google"] = True
                if gr.get("website") and not result.get("website"):
                    result["website"] = gr["website"]
                if gr.get("address") and not result.get("address"):
                    result["address"] = gr["address"]
                if gr.get("business_phone") and not result.get("business_phone"):
                    result["business_phone"] = gr["business_phone"]
            except Exception as e:
                _elog.warning("[enrich] Google Maps error for %s: %s", name, e)
        else:
            sources["google"] = True

        url = result.get("website", "")

        # 2. Sunbiz — try multiple name variants
        try:
            sb = sunbiz_lookup(name)
            if sb.get("found"):
                sources["sunbiz"] = True
                result["entity_name"]       = sb.get("entity_name", "") or result.get("entity_name", "")
                result["formation_date"]    = sb.get("date_filed", "") or result.get("formation_date", "")
                result["years_in_business"] = sb.get("years_in_business", "") or result.get("years_in_business", "")
                result["sunbiz_status"]     = sb.get("sunbiz_status", "") or result.get("sunbiz_status", "")
                result["sunbiz_url"]        = sb.get("sunbiz_url", "") or result.get("sunbiz_url", "")
                result["registered_agent"]  = sb.get("registered_agent", "") or result.get("registered_agent", "")
                result["reg_agent_address"] = sb.get("reg_agent_address", "") or result.get("reg_agent_address", "")
                if sb.get("owner_name") and not result.get("owner_name"):
                    result["owner_name"] = sb.get("owner_name", "")
            else:
                _elog.warning("[enrich] Sunbiz no match for %r (tried %s)", name, sb.get("tried"))
        except Exception as e:
            _elog.warning("[enrich] Sunbiz error for %s: %s", name, e)

        # 3. Website scrape — general email, Instagram, Facebook, phone
        if url:
            try:
                ws = scrape_website_contact(url)
                if ws.get("general_email"):
                    result["general_email"] = ws["general_email"]
                    sources["website"] = True
                if ws.get("instagram_url"):
                    result["instagram_url"] = ws["instagram_url"]
                    sources["website"] = True
                if ws.get("facebook_url") and not result.get("facebook_url"):
                    result["facebook_url"] = ws["facebook_url"]
                    sources["website"] = True
                if not result.get("business_phone") and ws.get("phones"):
                    result["business_phone"] = ws["phones"][0]
                    sources["website"] = True
            except Exception as e:
                _elog.warning("[enrich] Website scrape error for %s: %s", name, e)

        # 3b. Hunter.io — fill general_email if still blank
        if url and not result.get("general_email"):
            try:
                hunter_email = _hunter_domain_search(url)
                if hunter_email:
                    result["general_email"] = hunter_email
            except Exception as e:
                _elog.warning("[enrich] Hunter error for %s: %s", name, e)

        # 3c. Search-engine website fallback (no API key needed)
        if not result.get("website"):
            try:
                discovered = _discover_website_via_search(name, city, state)
                if discovered:
                    result["website"] = discovered
            except Exception as e:
                _elog.warning("[enrich] Website search fallback error for %s: %s", name, e)

        # 3d. Apollo organization fallback — final attempt to find a website
        if not result.get("website"):
            try:
                apollo_key = os.environ.get("APOLLO_API_KEY", "").strip()
                if apollo_key:
                    loc_terms = [t for t in [city, state] if t]
                    apollo_res = apollo_search_people(
                        keywords=name,
                        locations=loc_terms or None,
                        num_results=3,
                        _apollo_key=apollo_key,
                    )
                    orgs = (apollo_res or {}).get("organizations") or []
                    for org in orgs:
                        org_name = (org.get("name") or "").lower()
                        if not org_name:
                            continue
                        if name.lower().split()[0] in org_name or org_name.split()[0] in name.lower():
                            site = (org.get("website_url") or org.get("primary_domain") or "").strip()
                            if site:
                                if not site.startswith("http"):
                                    site = "https://" + site
                                result["website"] = site.rstrip("/")
                                sources["apollo"] = True
                            break
            except Exception as e:
                _elog.warning("[enrich] Apollo lookup error for %s: %s", name, e)

        # 3e. Re-scrape any newly discovered website for email/social/phone
        if result.get("website") and not result.get("general_email"):
            try:
                ws = scrape_website_contact(result["website"])
                if ws.get("general_email"):
                    result["general_email"] = ws["general_email"]
                    sources["website"] = True
                if ws.get("instagram_url") and not result.get("instagram_url"):
                    result["instagram_url"] = ws["instagram_url"]
                if ws.get("facebook_url") and not result.get("facebook_url"):
                    result["facebook_url"] = ws["facebook_url"]
                if not result.get("business_phone") and ws.get("phones"):
                    result["business_phone"] = ws["phones"][0]
            except Exception as e:
                _elog.warning("[enrich] Re-scrape error for %s: %s", name, e)

        # If Apollo gave us a website, scrape it now for email/social
        if sources["apollo"] and result.get("website") and not result.get("general_email"):
            try:
                ws = scrape_website_contact(result["website"])
                if ws.get("general_email"):
                    result["general_email"] = ws["general_email"]
                if ws.get("instagram_url") and not result.get("instagram_url"):
                    result["instagram_url"] = ws["instagram_url"]
                if ws.get("facebook_url") and not result.get("facebook_url"):
                    result["facebook_url"] = ws["facebook_url"]
            except Exception as e:
                _elog.warning("[enrich] Apollo website scrape error for %s: %s", name, e)

        # 4. Owner contact — email + cell phone via web search
        owner = result.get("owner_name", "")
        if owner and not (result.get("owner_email") and result.get("owner_phone")):
            try:
                oc = _find_person_contact(owner, name, city, state)
                if oc.get("email") and not result.get("owner_email"):
                    result["owner_email"] = oc["email"]
                if oc.get("phone") and not result.get("owner_phone"):
                    result["owner_phone"] = oc["phone"]
            except Exception as e:
                _elog.warning("[enrich] Owner contact error for %s: %s", name, e)

        # 5. Registered agent contact — email + cell phone via web search
        agent = result.get("registered_agent", "")
        if agent and not (result.get("reg_agent_email") and result.get("reg_agent_phone")):
            try:
                ac = _find_person_contact(agent, name, city, state)
                if ac.get("email") and not result.get("reg_agent_email"):
                    result["reg_agent_email"] = ac["email"]
                if ac.get("phone") and not result.get("reg_agent_phone"):
                    result["reg_agent_phone"] = ac["phone"]
            except Exception as e:
                _elog.warning("[enrich] Agent contact error for %s: %s", name, e)

        # If no email found anywhere, generate a placeholder from trade name
        if not result.get("general_email") and not result.get("owner_email"):
            slug = re.sub(r"[^a-z0-9]", "", (result.get("trade_name") or "business").lower())[:20]
            result["general_email"] = f"info@{slug}.com"

        return result

    # Each lead launches several Playwright browsers (Maps + maybe website search +
    # Sunbiz). On low-memory hosts (e.g. Render free) running these in parallel
    # OOM-kills Chromium silently and every Maps result returns blank. Run
    # sequentially in production, parallel only when explicitly enabled.
    parallel = os.environ.get("ENRICH_PARALLEL", "1") == "1" and not os.environ.get("RENDER")
    if parallel:
        with concurrent.futures.ThreadPoolExecutor(max_workers=2) as executor:
            futures = {executor.submit(_enrich_one, lead): i for i, lead in enumerate(leads)}
            for future in concurrent.futures.as_completed(futures):
                try:
                    enriched.append(future.result())
                except Exception as e:
                    _elog.warning("[enrich] worker failed: %s", e)
    else:
        for lead in leads:
            try:
                enriched.append(_enrich_one(lead))
            except Exception as e:
                _elog.warning("[enrich] sequential worker failed: %s", e)
                enriched.append(lead)

    enriched.sort(key=lambda x: x.get("trade_name", ""))

    sunbiz_hits  = sum(1 for l in enriched if l.get("sunbiz_url") or l.get("entity_name"))
    google_hits  = sum(1 for l in enriched if l.get("google_rating") or l.get("google_review_count"))
    website_hits = sum(1 for l in enriched if l.get("website"))
    email_hits   = sum(1 for l in enriched if l.get("general_email") and "@" in l.get("general_email", ""))

    _leads_store = enriched
    _save_leads_to_file(enriched)

    return {
        "leads": enriched,
        "total": len(enriched),
        "saved": True,
        "summary": {
            "sunbiz_matched":  sunbiz_hits,
            "google_matched":  google_hits,
            "websites_found":  website_hits,
            "emails_found":    email_hits,
            "out_of":          len(enriched),
        },
    }


TOOL_MAP = {
    "search_businesses_maps": search_businesses_maps,
    "web_search":             web_search,
    "apollo_search_people":   apollo_search_people,
    "apollo_enrich_existing_leads": apollo_enrich_existing_leads,
    "enrich_leads_batch":     enrich_leads_batch,
    "get_collected_leads":      get_collected_leads,
    "query_tenant_crm":         query_tenant_crm,
    "upload_leads_to_hubspot":  upload_leads_to_hubspot,
    "sunbiz_lookup":            sunbiz_lookup,
    "scrape_website_contact": scrape_website_contact,
    "get_google_reviews":     get_google_reviews,
    "hubspot_create_contact": hubspot_create_contact,
    "save_leads_csv":         save_leads_csv,
    "save_outreach_csv":      save_outreach_csv,
    "save_research_report":   save_research_report,
    "send_gmail_email":       send_gmail_email,
    "create_gmail_drafts":    create_gmail_drafts,
}


def run_tool(name, inputs, apollo_key="", hubspot_token=""):
    fn = TOOL_MAP.get(name)
    if fn is None:
        return {"error": f"Unknown tool: {name}"}
    kwargs = dict(inputs)
    if name in ("apollo_search_people", "apollo_enrich_existing_leads"):
        kwargs["_apollo_key"] = apollo_key
    elif name in ("hubspot_create_contact", "upload_leads_to_hubspot"):
        kwargs["_hubspot_token"] = hubspot_token
    return fn(**kwargs)


# ── Agentic loop (generator) ──────────────────────────────────────────────────

SYSTEM_PROMPT = """You are MMG Agent, a lead generation assistant for MMG — a commercial real estate brokerage that helps businesses find and lease commercial spaces.

## Your purpose
Find business prospects (tenants) who may be looking to open a new location, expand, or relocate — and help the user work them through discovery, enrichment, and HubSpot in a natural, step-by-step chat.

## Conversation flow (match the user's pace — one major action per turn unless they ask for several)

Follow this pattern when the user is building a pipeline (adapt wording to their request):

1. **Find leads** — Call `query_tenant_crm` first (local Miami-Dade / Broward database). Honor exact counts: if they say "5 nail salons", use `limit=5` and keywords like nail salon + Miami-Dade. Only if the database returns zero rows, use `search_businesses_maps` + `enrich_leads_batch`.
2. **Fill missing emails (Apollo)** — When they ask to "search Apollo", "find missing emails on Apollo", or "enrich with Apollo" for the **leads they already have**, call `apollo_enrich_existing_leads` (no arguments). It looks up each session lead on Apollo and fills in blank email/website/phone/industry fields without replacing the working set. Only call `apollo_search_people` if the user explicitly asks for **brand-new** Apollo leads (it REPLACES the working set).
3. **Show everything together** — When they want one combined view, call `get_collected_leads` and/or present the merged table from the current saved leads.
4. **Upload to HubSpot** — When they ask to push prospects to HubSpot, call `upload_leads_to_hubspot` (caps at 20 per call). Summarize uploaded vs skipped.
5. **Full research / Sunbiz / missing fields** — When they want a deep-dive report (Sunbiz, website, reviews, etc.), first call `enrich_leads_batch` on the current leads to refresh the data (no separate one-off sunbiz calls). Then write an in-depth markdown research report (see "Research report format" below) and call `save_research_report` with the full markdown so the user can download it.
6. **Update HubSpot after enrichment** — After enrichment, refresh HubSpot by calling `hubspot_create_contact` once per lead with the **same** email and updated `first_name`, `last_name`, `company`, `phone`, `website`, `linkedin` so records reflect new data (handle "already exists" gracefully in your reply).

## Required fields — for full enrichment pulls (enrich_leads_batch)

1.  Business trade name
2.  Business entity / corporate name (from Sunbiz)
3.  Company formation date + years in business
4.  Business general email
5.  Owner name
6.  Owner email
7.  Owner cell phone
8.  Registered Agent name
9.  Registered Agent email
10. Registered Agent cell phone
11. Business address
12. Business phone
13. Website URL
14. Instagram URL + Facebook URL
15. Google rating + Google review count

## Writing outreach emails
When asked to draft emails, call `save_outreach_csv` with personalized copy. Sign off as: MMG Real Estate Team.

## Research report format
When the user asks for a "research report", "deep dive", "full profile", etc., produce a markdown report covering ALL leads in the working set. Use only the data each lead actually has. NEVER fabricate.

**OMIT EMPTY DATA — do not show placeholders or "—".**
- For per-lead **field tables**: leave a row out entirely if there is no value. Do not write rows like `Instagram | —`.
- For per-lead **Ownership & Contacts tables**: leave a column out entirely if no row has a value for it (e.g. drop the Email column when no officer/agent has an email). Drop rows that have nothing but a name.
- For the **Summary table**: drop a column (e.g. "Years in Business") if no lead has data for it. Don't print "—" cells.
- Skip whole sections (e.g. "Notes", "Ownership & Contacts") if there is nothing meaningful to say. Do not write "No notes available."

Structure (adapt to what's available):

1. **Title block** — category, market, methodology (Sunbiz, Google Maps / Places API, websites, social, directories), brief executive summary (3–5 sentences) explaining selection criteria.
2. **Summary table** — one row per lead with rank, business, area, plus only the columns where at least one lead has data (e.g. Google rating, reviews, years in business, Sunbiz status).
3. **Per-lead profile** (`## N. Trade Name`):
   - One-sentence positioning headline.
   - **Field table** with only the rows that have values (Trade Name, Corporate Entity, Sunbiz Doc #, Sunbiz Status, Formation Date, Years in Business, Business Email, Business Phone, Business Address, Website, Instagram, Facebook, Google Rating + reviews).
   - **Ownership & Contacts table** if at least one role has data (Role, Name, then Email and/or Phone columns only if at least one row has them).
   - **Notes** paragraph only if there's something noteworthy (Sunbiz status changes, multi-entity owners, brand history, recent transitions, Maps attributes like women-owned).
   - **Prospecting Notes** paragraph: why this is a strong MMG prospect and the best contact channel.
4. **Outreach Priority Matrix** — table ranking leads (Highest / High / Medium / Low) with "why" and best contact (skip rows with no usable contact info).
5. **Research Notes & Disclaimers** — short notes on data sources, contact policy ("only publicly available info; nothing fabricated"), Sunbiz status key, review-count caveat, outreach compliance reminder.

After printing the report inline, call `save_research_report({report_markdown, title})` so it's downloadable. Reply with one short confirmation sentence after the tool call (e.g. "Saved — use the Download PDF button below.").

## Rules
- Keep replies short (1–2 sentences) after tool runs unless the user asked for a detailed report — then summarize findings clearly without dumping raw JSON.
- Do not use `web_search` unless the user explicitly asks.
- No markdown tables in chat — the UI shows tables when tools return leads.
- After `enrich_leads_batch`, ALWAYS report the per-source counts from the tool's `summary` object (e.g. "Sunbiz matched 4/5, Google Maps matched 5/5, websites 3/5"). NEVER claim "no Sunbiz, website, or Google data found" if any of those counts is greater than zero. If a count is zero for a category, say only that category was missing — don't generalize.
- If `enrich_leads_batch` returns 0 for a category but the user knows the data exists, suggest re-running enrichment (Sunbiz can rate-limit) before claiming the source has nothing.
"""


# ── Convert TOOLS → OpenAI / Gemini format ───────────────────────────────────

def _tools_openai_fmt():
    """Reformat TOOLS list from Anthropic schema to OpenAI function-calling schema."""
    return [
        {
            "type": "function",
            "function": {
                "name":        t["name"],
                "description": t["description"],
                "parameters":  t["input_schema"],
            },
        }
        for t in TOOLS
    ]


# ── Generic OpenAI-compatible agent loop ─────────────────────────────────────

def _run_agent_openai_compat(user_message: str, history: list,
                              api_key: str, model: str, base_url: str,
                              provider: str = "gemini",
                              apollo_key="", hubspot_token=""):
    """Shared agent loop for any OpenAI-compatible endpoint (Gemini, etc.)."""
    import time as _time
    from openai import OpenAI as _OAI

    client    = _OAI(api_key=api_key, base_url=base_url)
    oai_tools = _tools_openai_fmt()

    messages = [{"role": "system", "content": SYSTEM_PROMPT}]
    for msg in history:
        role    = msg.get("role")
        content = msg.get("content")
        if role in ("user", "assistant") and content:
            messages.append({"role": role, "content": content})
    messages.append({"role": "user", "content": user_message})

    t0 = _time.time()
    success = False
    total_in = total_out = total_tools = total_leads = 0
    try:
        while True:
            response = client.chat.completions.create(
                model=model, messages=messages,
                tools=oai_tools, tool_choice="auto",
            )
            usage = getattr(response, "usage", None)
            if usage:
                total_in  += getattr(usage, "prompt_tokens",     0) or 0
                total_out += getattr(usage, "completion_tokens", 0) or 0

            choice  = response.choices[0]
            msg_obj = choice.message

            if msg_obj.content:
                yield f"data: {json.dumps({'type': 'text', 'content': msg_obj.content})}\n\n"

            if choice.finish_reason == "stop" or not msg_obj.tool_calls:
                break

            messages.append(msg_obj)
            total_tools += len(msg_obj.tool_calls)

            tool_results = []
            for tc in msg_obj.tool_calls:
                name = tc.function.name
                try:
                    inputs = json.loads(tc.function.arguments)
                except Exception:
                    inputs = {}

                yield f"data: {json.dumps({'type': 'tool_start', 'name': name})}\n\n"
                result = run_tool(name, inputs,
                                  apollo_key=apollo_key,
                                  hubspot_token=hubspot_token)
                yield f"data: {json.dumps({'type': 'tool_end', 'name': name, 'result': result})}\n\n"
                if isinstance(result, dict) and result.get("leads"):
                    total_leads += len(result["leads"])

                tool_results.append({
                    "role":         "tool",
                    "tool_call_id": tc.id,
                    "content":      json.dumps(result),
                })

            messages.extend(tool_results)
        success = True

    except Exception as e:
        yield f"data: {json.dumps({'type': 'error', 'content': str(e)})}\n\n"
    finally:
        _record_perf(provider, model,
                     int((_time.time() - t0) * 1000),
                     total_in, total_out, total_tools, total_leads, success)

    yield f"data: {json.dumps({'type': 'done'})}\n\n"


def run_agent_gemini(user_message, history, gemini_key, model,
                     apollo_key="", hubspot_token=""):
    yield from _run_agent_openai_compat(
        user_message, history,
        api_key=gemini_key,
        model=model,
        base_url="https://generativelanguage.googleapis.com/v1beta/openai/",
        provider="gemini",
        apollo_key=apollo_key,
        hubspot_token=hubspot_token,
    )


def run_agent_perplexity(user_message, history, perplexity_key, model,
                         apollo_key="", hubspot_token=""):
    """
    Perplexity sonar models have built-in web search but do NOT support
    function/tool calling — so we use a plain chat completion loop.
    """
    import time as _time
    from openai import OpenAI as _OAI

    client = _OAI(api_key=perplexity_key, base_url="https://api.perplexity.ai/")

    messages = [{"role": "system", "content": SYSTEM_PROMPT}]
    for msg in history:
        role    = msg.get("role")
        content = msg.get("content")
        if role in ("user", "assistant") and content:
            messages.append({"role": role, "content": content})
    messages.append({"role": "user", "content": user_message})

    t0 = _time.time()
    success = False
    input_tokens = output_tokens = 0
    try:
        response = client.chat.completions.create(model=model, messages=messages)
        usage = getattr(response, "usage", None)
        if usage:
            input_tokens  = getattr(usage, "prompt_tokens",     0) or 0
            output_tokens = getattr(usage, "completion_tokens", 0) or 0
        text = response.choices[0].message.content or ""
        if text:
            yield f"data: {json.dumps({'type': 'text', 'content': text})}\n\n"
        success = True
    except Exception as e:
        yield f"data: {json.dumps({'type': 'error', 'content': str(e)})}\n\n"
    finally:
        _record_perf("perplexity", model,
                     int((_time.time() - t0) * 1000),
                     input_tokens, output_tokens, 0, 0, success)

    yield f"data: {json.dumps({'type': 'done'})}\n\n"


# ── Anthropic agent loop ──────────────────────────────────────────────────────

def run_agent_anthropic(user_message: str, history: list,
                        anthropic_key: str,
                        model="claude-opus-4-6",
                        apollo_key="", hubspot_token=""):
    """Agent loop using the Anthropic SDK."""
    import time as _time
    client = anthropic.Anthropic(api_key=anthropic_key)
    MODEL  = model or "claude-opus-4-6"

    api_messages = []
    for msg in history:
        role    = msg.get("role")
        content = msg.get("content")
        if role in ("user", "assistant") and content:
            api_messages.append({"role": role, "content": content})
    api_messages.append({"role": "user", "content": user_message})

    t0 = _time.time()
    success = False
    total_in = total_out = total_tools = total_leads = 0
    try:
        while True:
            response = client.messages.create(
                model=MODEL, max_tokens=4096,
                system=SYSTEM_PROMPT, tools=TOOLS,
                messages=api_messages,
            )
            total_in  += getattr(response.usage, "input_tokens",  0) or 0
            total_out += getattr(response.usage, "output_tokens", 0) or 0

            full_text  = ""
            tool_calls = []
            for block in response.content:
                if block.type == "text":
                    full_text += block.text
                elif block.type == "tool_use":
                    tool_calls.append(block)

            if full_text:
                yield f"data: {json.dumps({'type': 'text', 'content': full_text})}\n\n"

            if response.stop_reason == "end_turn" or not tool_calls:
                break

            api_messages.append({"role": "assistant", "content": response.content})
            total_tools += len(tool_calls)

            tool_results = []
            for tc in tool_calls:
                yield f"data: {json.dumps({'type': 'tool_start', 'name': tc.name})}\n\n"
                result = run_tool(tc.name, tc.input,
                                  apollo_key=apollo_key,
                                  hubspot_token=hubspot_token)
                yield f"data: {json.dumps({'type': 'tool_end', 'name': tc.name, 'result': result})}\n\n"
                if isinstance(result, dict) and result.get("leads"):
                    total_leads += len(result["leads"])
                tool_results.append({
                    "type":        "tool_result",
                    "tool_use_id": tc.id,
                    "content":     json.dumps(result),
                })

            api_messages.append({"role": "user", "content": tool_results})
        success = True

    except Exception as e:
        yield f"data: {json.dumps({'type': 'error', 'content': str(e)})}\n\n"
    finally:
        _record_perf("anthropic", MODEL,
                     int((_time.time() - t0) * 1000),
                     total_in, total_out, total_tools, total_leads, success)

    yield f"data: {json.dumps({'type': 'done'})}\n\n"


# ── Unified entry point ───────────────────────────────────────────────────────

_LEAD_KEYWORDS = {
    "hair salon", "hair salons", "barbershop", "barbershops", "barber", "barbers",
    "nail salon", "nail salons", "salon", "salons", "beauty salon", "beauty salons",
    "show me", "find me", "pull", "list", "get me", "search for",
    "leads", "businesses", "contacts",
}

def _message_wants_leads(msg: str) -> bool:
    """Return True if the message is asking for businesses/leads from the local DB."""
    low = msg.lower()
    return any(kw in low for kw in _LEAD_KEYWORDS)


def _auto_crm_prefix(user_message: str) -> str:
    """
    If the message sounds like a lead request, call query_tenant_crm immediately
    and return a system note with the results to prepend to the conversation.
    Returns empty string if the DB is unavailable or returns no rows.
    """
    if not _message_wants_leads(user_message):
        return ""
    path = _leads_db_xlsx_path()
    if not path or not os.path.isfile(path):
        return ""
    # Extract keywords from the message to use as query
    low = user_message.lower()
    # Strip very generic words so we get meaningful search terms
    stop = {"show", "me", "find", "get", "list", "pull", "in", "from", "the",
            "a", "an", "for", "and", "or", "of", "all", "some", "please", "i", "want"}
    words = [w for w in re.sub(r"[^\w\s]", " ", low).split() if w not in stop]
    query = " ".join(words[:6])  # max 6 keywords
    try:
        wants_all = _wants_all_results(low)
        result = query_tenant_crm(query=query, limit=(_CRM_MAX_LIMIT if wants_all else 50), all_results=wants_all)
        if result.get("error") or not result.get("rows"):
            return ""
        rows = result["rows"]
        count = result["count"]
        # Format as a compact summary for the model
        lines = [f"[DATABASE RESULTS — {count} matches for '{query}' from local leads DB]"]
        for r in rows[:50]:
            name = r.get("Business_Name", r.get("business_name", ""))
            addr = r.get("Business_Address", r.get("business_address", ""))
            phone = r.get("Business_Phone", r.get("business_phone", ""))
            email = r.get("Business_Email", r.get("business_email", ""))
            sheet = r.get("sheet", "")
            lines.append(f"- {name} | {addr} | {phone} | {email} | [{sheet}]")
        lines.append("[END DATABASE RESULTS — present these to the user, do NOT search Maps]")
        return "\n".join(lines)
    except Exception:
        return ""


def run_agent(user_message: str, history: list,
              anthropic_key="", apollo_key="", hubspot_token="",
              claude_model="claude-opus-4-6",
              gemini_key="", model_provider="anthropic",
              gemini_model="gemini-3-flash-preview",
              perplexity_key="", perplexity_model="sonar-pro"):
    """Route to the right model provider based on settings."""
    # Pre-query the local DB and inject results so the model can't skip them
    crm_prefix = _auto_crm_prefix(user_message)
    if crm_prefix:
        user_message = user_message + "\n\n" + crm_prefix

    if model_provider == "gemini":
        gemini_key = gemini_key or os.getenv("GEMINI_API_KEY", "")
        if not gemini_key:
            yield f"data: {json.dumps({'type': 'text', 'content': 'Please configure your Gemini API key in Settings.'})}\n\n"
            yield f"data: {json.dumps({'type': 'done'})}\n\n"
            return
        yield from run_agent_gemini(user_message, history,
                                    gemini_key=gemini_key,
                                    model=gemini_model,
                                    apollo_key=apollo_key,
                                    hubspot_token=hubspot_token)
        return

    if model_provider == "perplexity":
        perplexity_key = perplexity_key or os.getenv("PERPLEXITY_API_KEY", "")
        if not perplexity_key:
            yield f"data: {json.dumps({'type': 'text', 'content': 'Please configure your Perplexity API key in Settings.'})}\n\n"
            yield f"data: {json.dumps({'type': 'done'})}\n\n"
            return
        yield from run_agent_perplexity(user_message, history,
                                        perplexity_key=perplexity_key,
                                        model=perplexity_model,
                                        apollo_key=apollo_key,
                                        hubspot_token=hubspot_token)
        return

    # Default: Anthropic
    anthropic_key = anthropic_key or os.getenv("ANTHROPIC_API_KEY", "")
    if not anthropic_key:
        yield f"data: {json.dumps({'type': 'text', 'content': 'Please configure your Anthropic API key in Settings.'})}\n\n"
        yield f"data: {json.dumps({'type': 'done'})}\n\n"
        return
    yield from run_agent_anthropic(user_message, history,
                                   anthropic_key=anthropic_key,
                                   model=claude_model,
                                   apollo_key=apollo_key,
                                   hubspot_token=hubspot_token)


# ── Flask routes ──────────────────────────────────────────────────────────────

@app.after_request
def add_cors(response):
    response.headers["Access-Control-Allow-Origin"]  = "*"
    response.headers["Access-Control-Allow-Headers"] = "Content-Type"
    response.headers["Access-Control-Allow-Methods"] = "GET,POST,OPTIONS"
    return response


@app.route("/")
def index():
    return render_template("index.html")


@app.route("/api/diagnose")
def diagnose_enrichment():
    """
    Run a quick end-to-end check of every enrichment source so you can see
    exactly what's broken on Render vs local. No secrets are returned.

    Optional query params:
      ?biz=Joes+Stone+Crab&city=Miami+Beach&state=FL  (override the test target)
    """
    biz   = (request.args.get("biz")   or "Joe's Stone Crab").strip()
    city  = (request.args.get("city")  or "Miami Beach").strip()
    state = (request.args.get("state") or "FL").strip()

    out = {
        "test_business": f"{biz} ({city}, {state})",
        "host": {
            "render":             bool(os.environ.get("RENDER")),
            "playwright_browsers_path": os.environ.get("PLAYWRIGHT_BROWSERS_PATH") or "",
            "outbound_proxy":     bool(_outbound_proxy_url()),
            "google_places_key":  bool(
                (os.getenv("GOOGLE_PLACES_API_KEY") or os.getenv("GOOGLE_MAPS_API_KEY") or "").strip()
            ),
            "apollo_key":         bool(os.getenv("APOLLO_API_KEY", "").strip()),
            "hunter_key":         bool(os.getenv("HUNTER_API_KEY", "").strip()),
        },
        "checks": {},
    }

    # 1) Outbound IP (helpful: tells you what Google sees)
    try:
        r = requests.get("https://api.ipify.org?format=json", timeout=8,
                         proxies=_proxy_dict_for_requests())
        out["host"]["egress_ip"] = (r.json() or {}).get("ip", "")
    except Exception as e:
        out["host"]["egress_ip_error"] = str(e)

    # 2) Playwright launch test
    try:
        from playwright.sync_api import sync_playwright
        import time as _t
        t0 = _t.time()
        with sync_playwright() as pw:
            launch_kwargs = dict(
                headless=True,
                args=["--no-sandbox", "--disable-dev-shm-usage", "--disable-gpu",
                      "--disable-blink-features=AutomationControlled"],
            )
            proxy_cfg = _proxy_for_playwright()
            if proxy_cfg:
                launch_kwargs["proxy"] = proxy_cfg
            browser = pw.chromium.launch(**launch_kwargs)
            ctx = browser.new_context(
                user_agent="Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) "
                           "AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36",
                viewport={"width": 1366, "height": 900}, locale="en-US",
            )
            page = ctx.new_page()
            page.goto("https://www.google.com/search?q=hello", wait_until="domcontentloaded", timeout=20000)
            url = page.url
            html_len = len(page.content())
            browser.close()
        out["checks"]["playwright_google"] = {
            "ok":           True,
            "final_url":    url,
            "consent_wall": "consent.google.com" in url,
            "html_length":  html_len,
            "ms":           int((_t.time() - t0) * 1000),
        }
    except Exception as e:
        out["checks"]["playwright_google"] = {"ok": False, "error": str(e)}

    # 3) Google Places API
    try:
        pl = _google_places_lookup(biz, city, state)
        out["checks"]["google_places_api"] = {
            "configured": bool((os.getenv("GOOGLE_PLACES_API_KEY") or os.getenv("GOOGLE_MAPS_API_KEY") or "").strip()),
            "rating":     pl.get("google_rating", ""),
            "reviews":    pl.get("google_review_count", ""),
            "website":    pl.get("website", ""),
            "phone":      pl.get("business_phone", ""),
            "error":      pl.get("_error", ""),
        }
    except Exception as e:
        out["checks"]["google_places_api"] = {"ok": False, "error": str(e)}

    # 4) Headless Maps scrape
    try:
        gr = get_google_reviews(biz, city, state)
        out["checks"]["headless_maps"] = {
            "rating":  gr.get("google_rating", ""),
            "reviews": gr.get("google_review_count", ""),
            "website": gr.get("website", ""),
            "phone":   gr.get("business_phone", ""),
        }
    except Exception as e:
        out["checks"]["headless_maps"] = {"ok": False, "error": str(e)}

    # 5) Sunbiz (different host, different blocking pattern)
    try:
        sb = sunbiz_lookup(biz)
        out["checks"]["sunbiz"] = {
            "found":       bool(sb.get("found")),
            "entity_name": sb.get("entity_name", ""),
            "status":      sb.get("sunbiz_status", ""),
            "error":       sb.get("error", ""),
        }
    except Exception as e:
        out["checks"]["sunbiz"] = {"ok": False, "error": str(e)}

    # 6) Search-engine website discovery
    try:
        site = _discover_website_via_requests(f"{biz} {city} {state} official site")
        out["checks"]["website_search_requests"] = {"website": site}
    except Exception as e:
        out["checks"]["website_search_requests"] = {"ok": False, "error": str(e)}

    return jsonify(out)


@app.route("/api/config", methods=["GET"])
def get_config():
    _gmail_addr, _gmail_pw = _get_gmail_creds()
    gmail_connected = bool(_gmail_addr and _gmail_pw)
    crm_path = _setting_str("crm_path", (os.getenv("TENANT_CRM_XLSX_PATH") or "").strip())
    try:
        import dotenv as _dotenv_mod  # noqa: F401
        _dotenv_ok = True
    except ImportError:
        _dotenv_ok = False
    return jsonify({
        "anthropic":      bool(_credential("anthropic_key", "ANTHROPIC_API_KEY")),
        "apollo":         bool(_credential("apollo_key", "APOLLO_API_KEY")),
        "hubspot":        bool(_credential("hubspot_token", "HUBSPOT_TOKEN")),
        "gemini":         bool(_credential("gemini_key", "GEMINI_API_KEY")),
        "perplexity":     bool(_credential("perplexity_key", "PERPLEXITY_API_KEY")),
        "gmail":          gmail_connected,
        "crm":            bool(crm_path) and os.path.isfile(os.path.expanduser(crm_path)),
        "model_provider":   _setting_str("model_provider", "anthropic"),
        "claude_model":     _setting_str("claude_model", "claude-opus-4-6"),
        "gemini_model":     _setting_str("gemini_model", "gemini-3-flash-preview"),
        "perplexity_model": _setting_str("perplexity_model", "sonar-pro"),
        # Local debugging: why keys are missing (no secret values)
        "env_file_loaded": bool(_DOTENV_LOADED_PATH),
        "env_file_path":   _DOTENV_LOADED_PATH or "",
        "python_dotenv_installed": _dotenv_ok,
        "flask_app_dir":   os.path.dirname(os.path.abspath(__file__)),
        # True when Google Places API key is set — recommended on Render (Maps scraping often blocked)
        "google_places":  bool(
            (os.getenv("GOOGLE_PLACES_API_KEY") or os.getenv("GOOGLE_MAPS_API_KEY") or "").strip()
        ),
    })


@app.route("/api/config", methods=["POST"])
def save_config():
    data = request.get_json(force=True) or {}
    local = _read_local_api_config()

    def _merge_setting(key: str):
        if key not in data:
            return
        raw = data[key]
        if raw is None:
            return
        if isinstance(raw, str):
            raw = raw.strip()
        # Empty password fields in the UI often mean "unchanged" — do not wipe saved keys
        if key in _LOCAL_SECRET_KEYS and raw == "":
            return
        if raw == "" or raw == []:
            local.pop(key, None)
            session.pop(key, None)
        else:
            local[key] = raw
            session[key] = raw

    for k in _LOCAL_CONFIG_KEYS:
        _merge_setting(k)

    global _hunter_key
    _hunter_key = _credential("hunter_key", "HUNTER_API_KEY")

    try:
        _write_local_api_config(local)
    except OSError as e:
        return jsonify({"ok": False, "error": f"Could not save settings: {e}"}), 500

    if data.get("gmail_address") or data.get("gmail_app_password"):
        # Persist to file so credentials survive server restarts
        _gmail_app_file = os.path.join(os.path.dirname(os.path.abspath(__file__)), ".gmail_app.json")
        existing = {}
        if os.path.exists(_gmail_app_file):
            try:
                with open(_gmail_app_file) as f:
                    existing = json.load(f)
            except Exception:
                pass
        if data.get("gmail_address"):
            existing["gmail_address"] = data["gmail_address"]
        if data.get("gmail_app_password"):
            existing["gmail_app_password"] = data["gmail_app_password"]
        with open(_gmail_app_file, "w") as f:
            json.dump(existing, f)
        session["gmail_address"]      = existing.get("gmail_address", "")
        session["gmail_app_password"] = existing.get("gmail_app_password", "")
    return jsonify({"ok": True})


@app.route("/api/chat", methods=["POST"])
def chat():
    data    = request.get_json(force=True)
    message = data.get("message", "")
    history = data.get("history", [])

    # Session + disk (.local_api_keys.json) + env
    anthropic_key    = _credential("anthropic_key", "ANTHROPIC_API_KEY")
    apollo_key       = _credential("apollo_key", "APOLLO_API_KEY")
    hubspot_token    = _credential("hubspot_token", "HUBSPOT_TOKEN")
    gemini_key       = _credential("gemini_key", "GEMINI_API_KEY")
    model_provider   = _setting_str("model_provider", "anthropic")
    claude_model     = _setting_str("claude_model", "claude-opus-4-6")
    gemini_model     = _setting_str("gemini_model", "gemini-2.0-flash")
    perplexity_key   = _credential("perplexity_key", "PERPLEXITY_API_KEY")
    global _hunter_key
    _hunter_key = _credential("hunter_key", "HUNTER_API_KEY") or _hunter_key
    perplexity_model = _setting_str("perplexity_model", "sonar-pro")

    def stream():
        try:
            yield from run_agent(message, history,
                                 anthropic_key=anthropic_key,
                                 apollo_key=apollo_key,
                                 hubspot_token=hubspot_token,
                                 claude_model=claude_model,
                                 gemini_key=gemini_key,
                                 model_provider=model_provider,
                                 gemini_model=gemini_model,
                                 perplexity_key=perplexity_key,
                                 perplexity_model=perplexity_model)
        except Exception as e:
            app.logger.error("Unhandled stream error [%s]: %s", model_provider, e, exc_info=True)
            yield f"data: {json.dumps({'type': 'error', 'content': str(e)})}\n\n"
            yield f"data: {json.dumps({'type': 'done'})}\n\n"

    return Response(
        stream(),
        mimetype="text/event-stream",
        headers={"Cache-Control": "no-cache", "X-Accel-Buffering": "no"},
    )


@app.route("/api/performance")
def get_performance():
    """Return aggregated and raw performance stats."""
    from collections import defaultdict

    records = list(_perf_store)  # snapshot

    # Per-provider aggregates
    agg = defaultdict(lambda: {
        "requests": 0, "successes": 0,
        "total_ms": 0, "total_in": 0, "total_out": 0,
        "total_tools": 0, "total_leads": 0, "total_cost": 0.0,
    })
    for r in records:
        p = r["provider"]
        agg[p]["requests"]    += 1
        agg[p]["successes"]   += 1 if r["success"] else 0
        agg[p]["total_ms"]    += r["duration_ms"]
        agg[p]["total_in"]    += r["input_tokens"]
        agg[p]["total_out"]   += r["output_tokens"]
        agg[p]["total_tools"] += r["tool_calls"]
        agg[p]["total_leads"] += r["leads_found"]
        agg[p]["total_cost"]  += r["cost_usd"]

    summary = {}
    for p, d in agg.items():
        n = d["requests"]
        summary[p] = {
            "requests":       n,
            "success_rate":   round(d["successes"] / n * 100, 1) if n else 0,
            "avg_ms":         round(d["total_ms"] / n) if n else 0,
            "total_tokens":   d["total_in"] + d["total_out"],
            "avg_tokens":     round((d["total_in"] + d["total_out"]) / n) if n else 0,
            "total_leads":    d["total_leads"],
            "avg_leads":      round(d["total_leads"] / n, 1) if n else 0,
            "total_cost_usd": round(d["total_cost"], 4),
            "avg_cost_usd":   round(d["total_cost"] / n, 4) if n else 0,
        }

    # Last 20 raw records (newest first)
    recent = sorted(records, key=lambda r: r["ts"], reverse=True)[:20]

    return jsonify({"summary": summary, "recent": recent})


@app.route("/api/download/leads")
def download_leads():
    path = os.path.join(os.path.dirname(__file__), "leads.csv")
    if os.path.exists(path):
        return send_file(path, mimetype="text/csv",
                         as_attachment=True, download_name="leads.csv")
    global _leads_store
    if not _leads_store:
        return jsonify({"error": "No leads available"}), 404
    output = io.StringIO()
    writer = csv.DictWriter(output, fieldnames=LEAD_FIELDS, extrasaction="ignore")
    writer.writeheader()
    writer.writerows(_leads_store)
    return Response(
        output.getvalue(),
        mimetype="text/csv",
        headers={"Content-Disposition": "attachment; filename=leads.csv"},
    )


@app.route("/api/save_outreach", methods=["POST"])
def save_outreach_edits():
    """Save inline-edited outreach drafts back to server."""
    global _outreach_store
    data = request.get_json(force=True)
    drafts = data.get("drafts", [])
    _outreach_store = drafts
    try:
        path = os.path.join(os.path.dirname(__file__), "outreach_drafts.csv")
        with open(path, "w", newline="") as f:
            writer = csv.DictWriter(
                f, fieldnames=["name", "email", "subject_line", "email_body"],
                extrasaction="ignore",
            )
            writer.writeheader()
            writer.writerows(drafts)
        return jsonify({"success": True, "count": len(drafts)})
    except Exception as e:
        return jsonify({"error": str(e)}), 500


@app.route("/api/download/outreach")
def download_outreach():
    path = os.path.join(os.path.dirname(__file__), "outreach_drafts.csv")
    if os.path.exists(path):
        return send_file(path, mimetype="text/csv",
                         as_attachment=True, download_name="outreach_drafts.csv")
    global _outreach_store
    if not _outreach_store:
        return jsonify({"error": "No outreach drafts available"}), 404
    output = io.StringIO()
    writer = csv.DictWriter(
        output, fieldnames=["name", "email", "subject_line", "email_body"],
        extrasaction="ignore",
    )
    writer.writeheader()
    writer.writerows(_outreach_store)
    return Response(
        output.getvalue(),
        mimetype="text/csv",
        headers={"Content-Disposition": "attachment; filename=outreach_drafts.csv"},
    )


@app.route("/api/clear_leads", methods=["POST"])
def clear_leads():
    global _leads_store, _outreach_store
    _leads_store = []
    _outreach_store = []
    return jsonify({"ok": True})


def _research_report_html(title: str, markdown_body: str) -> str:
    """Wrap the report markdown in a print-friendly HTML page."""
    safe_title = (title or "Tenant Prospecting Research Report").replace("<", "&lt;").replace(">", "&gt;")
    return (
        "<!DOCTYPE html><html lang=\"en\"><head>"
        f"<meta charset=\"utf-8\"><title>{safe_title}</title>"
        "<script src=\"https://cdn.jsdelivr.net/npm/marked/marked.min.js\"></script>"
        "<style>"
        "body{font-family:-apple-system,Segoe UI,Inter,Helvetica,Arial,sans-serif;color:#111;"
        "max-width:920px;margin:40px auto;padding:0 24px;line-height:1.55;}"
        "h1{font-size:1.6rem;border-bottom:2px solid #0f766e;padding-bottom:.4rem;}"
        "h2{font-size:1.2rem;margin-top:1.6rem;color:#0f766e;}"
        "h3{font-size:1.05rem;margin-top:1.1rem;}"
        "table{width:100%;border-collapse:collapse;margin:.6rem 0;font-size:.92rem;}"
        "th,td{border:1px solid #e5e7eb;padding:.45rem .6rem;text-align:left;vertical-align:top;}"
        "th{background:#f3f4f6;font-weight:600;}"
        "blockquote{border-left:3px solid #0f766e;color:#374151;margin:.8rem 0;padding:.2rem .8rem;background:#f9fafb;}"
        "code{background:#f3f4f6;padding:.1em .35em;border-radius:4px;font-family:ui-monospace,monospace;}"
        "@media print{a{color:inherit;text-decoration:none;}body{margin:0.4in;}}"
        ".meta{color:#6b7280;font-size:.85rem;margin-top:-.4rem;}"
        ".actions{margin:.5rem 0 1.4rem;}"
        ".actions button{background:#0f766e;color:#fff;border:0;padding:.4rem .8rem;border-radius:6px;cursor:pointer;}"
        "</style></head><body>"
        f"<h1>{safe_title}</h1>"
        "<p class=\"meta\">Generated by MMG Agent</p>"
        "<div class=\"actions\"><button onclick=\"window.print()\">Print or Save as PDF</button></div>"
        f"<div id=\"report-body\" data-md=\"{markdown_body and ''}\"></div>"
        "<script>"
        f"const _md = {json.dumps(markdown_body)};"
        "document.getElementById('report-body').innerHTML = marked.parse(_md);"
        "</script>"
        "</body></html>"
    )


@app.route("/api/report")
def view_research_report():
    if not _research_report_store.get("markdown"):
        return "<p style=\"font-family:sans-serif;padding:2rem\">No research report yet — ask the agent for a research report.</p>", 404
    return _research_report_print_html(_research_report_store.get("title", ""), _research_report_store["markdown"])


def _markdown_to_basic_html(md: str) -> str:
    """Render a small subset of markdown to HTML server-side (so PDF generators
    don't need JavaScript). Supports # / ## / ### headings, bold, italic, links,
    fenced code, blockquotes, GitHub-style tables, and bullet/numbered lists."""
    lines = (md or "").splitlines()
    out = []
    i = 0

    def _inline(s: str) -> str:
        s = (s.replace("&", "&amp;")
              .replace("<", "&lt;")
              .replace(">", "&gt;"))
        s = re.sub(r"`([^`]+)`", r"<code>\1</code>", s)
        s = re.sub(r"\*\*([^*]+)\*\*", r"<strong>\1</strong>", s)
        s = re.sub(r"(?<!\*)\*([^*\n]+)\*", r"<em>\1</em>", s)
        s = re.sub(r"\[([^\]]+)\]\((https?://[^)]+)\)", r'<a href="\2">\1</a>', s)
        return s

    while i < len(lines):
        line = lines[i]
        stripped = line.strip()

        if not stripped:
            out.append("")
            i += 1
            continue

        if stripped.startswith("```"):
            i += 1
            buf = []
            while i < len(lines) and not lines[i].startswith("```"):
                buf.append(lines[i])
                i += 1
            i += 1
            out.append("<pre><code>" + "\n".join(
                l.replace("&", "&amp;").replace("<", "&lt;").replace(">", "&gt;") for l in buf
            ) + "</code></pre>")
            continue

        m = re.match(r"^(#{1,6})\s+(.*)", stripped)
        if m:
            level = len(m.group(1))
            out.append(f"<h{level}>{_inline(m.group(2).strip())}</h{level}>")
            i += 1
            continue

        if stripped.startswith(">"):
            buf = []
            while i < len(lines) and lines[i].strip().startswith(">"):
                buf.append(lines[i].strip().lstrip(">").strip())
                i += 1
            out.append("<blockquote>" + " ".join(_inline(b) for b in buf) + "</blockquote>")
            continue

        # GitHub-style table: header row, separator row of dashes, then body
        if "|" in stripped and i + 1 < len(lines) and re.match(r"^\s*\|?[\s\-:|]+\|?\s*$", lines[i + 1]) and "---" in lines[i + 1]:
            def _split_row(row):
                row = row.strip()
                if row.startswith("|"): row = row[1:]
                if row.endswith("|"):   row = row[:-1]
                return [c.strip() for c in row.split("|")]
            headers = _split_row(stripped)
            i += 2
            body = []
            while i < len(lines) and lines[i].strip() and "|" in lines[i]:
                body.append(_split_row(lines[i]))
                i += 1
            html = "<table><thead><tr>" + "".join(f"<th>{_inline(h)}</th>" for h in headers) + "</tr></thead><tbody>"
            for row in body:
                cells = (row + [""] * len(headers))[: len(headers)]
                html += "<tr>" + "".join(f"<td>{_inline(c)}</td>" for c in cells) + "</tr>"
            html += "</tbody></table>"
            out.append(html)
            continue

        # Unordered list
        if re.match(r"^\s*[-*+]\s+", line):
            items = []
            while i < len(lines) and re.match(r"^\s*[-*+]\s+", lines[i]):
                items.append(re.sub(r"^\s*[-*+]\s+", "", lines[i]))
                i += 1
            out.append("<ul>" + "".join(f"<li>{_inline(x)}</li>" for x in items) + "</ul>")
            continue

        # Ordered list
        if re.match(r"^\s*\d+\.\s+", line):
            items = []
            while i < len(lines) and re.match(r"^\s*\d+\.\s+", lines[i]):
                items.append(re.sub(r"^\s*\d+\.\s+", "", lines[i]))
                i += 1
            out.append("<ol>" + "".join(f"<li>{_inline(x)}</li>" for x in items) + "</ol>")
            continue

        # Paragraph (collapse adjacent non-block lines)
        buf = [stripped]
        i += 1
        while i < len(lines):
            nxt = lines[i].strip()
            if (not nxt) or re.match(r"^(#{1,6}\s|>|[-*+]\s|\d+\.\s|```)", nxt) or "|" in nxt:
                break
            buf.append(nxt)
            i += 1
        out.append("<p>" + _inline(" ".join(buf)) + "</p>")

    return "\n".join(out)


def _research_report_print_html(title: str, markdown_body: str) -> str:
    """Print-quality HTML rendered server-side (no client JS), used for PDF."""
    safe_title = (title or "Tenant Prospecting Research Report").replace("<", "&lt;").replace(">", "&gt;")
    body_html = _markdown_to_basic_html(markdown_body or "")
    return (
        "<!DOCTYPE html><html lang=\"en\"><head><meta charset=\"utf-8\">"
        f"<title>{safe_title}</title>"
        "<style>"
        "@page { size: Letter; margin: 0.55in; }"
        "body{font-family:-apple-system,Segoe UI,Inter,Helvetica,Arial,sans-serif;color:#111;"
        "max-width:920px;margin:0 auto;line-height:1.45;font-size:11pt;}"
        "h1{font-size:18pt;border-bottom:2px solid #0f766e;padding-bottom:.3rem;margin-top:0;}"
        "h2{font-size:13pt;color:#0f766e;margin-top:1.2rem;page-break-after:avoid;}"
        "h3{font-size:11.5pt;margin-top:.9rem;page-break-after:avoid;}"
        "table{width:100%;border-collapse:collapse;margin:.5rem 0;font-size:10pt;page-break-inside:avoid;}"
        "th,td{border:1px solid #d1d5db;padding:.35rem .55rem;text-align:left;vertical-align:top;}"
        "th{background:#f3f4f6;font-weight:600;}"
        "blockquote{border-left:3px solid #0f766e;color:#374151;margin:.7rem 0;padding:.15rem .7rem;background:#f9fafb;}"
        "code{background:#f3f4f6;padding:.05em .35em;border-radius:3px;font-family:ui-monospace,monospace;font-size:.9em;}"
        "ul,ol{margin:.4rem 0 .6rem 1.2rem;}"
        "p{margin:.45rem 0;}"
        ".meta{color:#6b7280;font-size:9pt;margin-top:-.3rem;}"
        "</style></head><body>"
        f"<h1>{safe_title}</h1>"
        "<p class=\"meta\">Generated by MMG Agent</p>"
        f"{body_html}"
        "</body></html>"
    )


def _render_pdf_from_html(html: str) -> bytes:
    """Render HTML to PDF using whichever engine is available.

    Order:
      1. Playwright (Chromium) — already a project dep, identical Letter-size output
         on Render and local. Most reliable when cairo isn't available.
      2. WeasyPrint — if installed (needs cairo system libs).
      3. xhtml2pdf — pure-Python fallback (smaller, less polished tables).
    Returns b"" if all attempts fail."""
    import logging as _logging
    log = _logging.getLogger(__name__)

    # 1) Playwright print-to-PDF — same engine that prints from a real browser.
    try:
        from playwright.sync_api import sync_playwright
        launch_kwargs = dict(
            headless=True,
            args=["--no-sandbox", "--disable-dev-shm-usage", "--disable-gpu"],
        )
        with sync_playwright() as pw:
            browser = pw.chromium.launch(**launch_kwargs)
            ctx = browser.new_context()
            page = ctx.new_page()
            page.set_content(html, wait_until="domcontentloaded")
            pdf = page.pdf(
                format="Letter",
                margin={"top": "0.55in", "right": "0.55in", "bottom": "0.55in", "left": "0.55in"},
                print_background=True,
            )
            browser.close()
            if pdf:
                return pdf
    except Exception as e:
        log.warning("[pdf] playwright failed: %s", e)

    # 2) WeasyPrint — if system has cairo + pango
    try:
        from weasyprint import HTML  # type: ignore
        return HTML(string=html, base_url=request.host_url).write_pdf() or b""
    except Exception as e:
        log.warning("[pdf] weasyprint unavailable: %s", e)

    # 3) xhtml2pdf — pure-Python fallback
    try:
        from xhtml2pdf import pisa  # type: ignore
        buf = io.BytesIO()
        result = pisa.CreatePDF(src=html, dest=buf)
        if not result.err:
            return buf.getvalue()
    except Exception as e:
        log.warning("[pdf] xhtml2pdf unavailable: %s", e)

    return b""


@app.route("/api/download/report")
def download_research_report():
    if not _research_report_store.get("markdown"):
        return jsonify({"error": "No research report available"}), 404
    title = _research_report_store.get("title", "research_report")
    fname = re.sub(r"[^a-zA-Z0-9_\-]+", "_", title).strip("_") or "research_report"
    html  = _research_report_print_html(title, _research_report_store["markdown"])

    pdf_bytes = _render_pdf_from_html(html)
    if pdf_bytes:
        return Response(
            pdf_bytes,
            mimetype="application/pdf",
            headers={"Content-Disposition": f"attachment; filename={fname}.pdf"},
        )

    # Fallback: serve a print-ready HTML that auto-opens the print dialog
    print_html = html.replace(
        "</body></html>",
        "<script>window.addEventListener('load',()=>setTimeout(()=>window.print(),300));</script></body></html>",
    )
    return Response(
        print_html,
        mimetype="text/html",
        headers={"Content-Disposition": f"inline; filename={fname}.html"},
    )


# ── Gmail OAuth endpoints ──────────────────────────────────────────────────────

@app.route("/api/gmail/auth")
def gmail_auth():
    if not _GMAIL_AVAILABLE:
        return jsonify({"error": "Gmail libraries not installed"}), 500
    if not os.path.exists(_GMAIL_CLIENT_FILE):
        return jsonify({"error": "Gmail OAuth credentials not configured. Add Client ID + Secret in Settings first."}), 400
    with open(_GMAIL_CLIENT_FILE) as f:
        client_data = json.load(f)
    client_id     = client_data.get("client_id", "")
    client_secret = client_data.get("client_secret", "")
    if not client_id or not client_secret:
        return jsonify({"error": "Gmail Client ID or Secret is empty"}), 400

    flow = GoogleFlow.from_client_config(
        {
            "web": {
                "client_id":     client_id,
                "client_secret": client_secret,
                "auth_uri":      "https://accounts.google.com/o/oauth2/auth",
                "token_uri":     "https://oauth2.googleapis.com/token",
                "redirect_uris": [_GMAIL_REDIRECT_URI],
            }
        },
        scopes=_GMAIL_SCOPES,
    )
    flow.redirect_uri = _GMAIL_REDIRECT_URI
    authorization_url, state = flow.authorization_url(
        access_type="offline",
        include_granted_scopes="true",
        prompt="consent",
    )
    session["gmail_oauth_state"] = state
    return redirect(authorization_url)


@app.route("/api/gmail/callback")
def gmail_callback():
    if not _GMAIL_AVAILABLE:
        return "<p>Gmail libraries not installed.</p>", 500
    if not os.path.exists(_GMAIL_CLIENT_FILE):
        return "<p>Gmail client credentials missing.</p>", 400
    with open(_GMAIL_CLIENT_FILE) as f:
        client_data = json.load(f)
    client_id     = client_data.get("client_id", "")
    client_secret = client_data.get("client_secret", "")

    flow = GoogleFlow.from_client_config(
        {
            "web": {
                "client_id":     client_id,
                "client_secret": client_secret,
                "auth_uri":      "https://accounts.google.com/o/oauth2/auth",
                "token_uri":     "https://oauth2.googleapis.com/token",
                "redirect_uris": [_GMAIL_REDIRECT_URI],
            }
        },
        scopes=_GMAIL_SCOPES,
        state=session.get("gmail_oauth_state"),
    )
    flow.redirect_uri = _GMAIL_REDIRECT_URI

    try:
        # Allow HTTP for local dev
        import os as _os
        _os.environ["OAUTHLIB_INSECURE_TRANSPORT"] = "1"
        flow.fetch_token(authorization_response=request.url)
        creds = flow.credentials
        _save_gmail_creds(creds, client_id, client_secret)
        return """
        <html><body style="font-family:sans-serif;display:flex;align-items:center;justify-content:center;height:100vh;margin:0;background:#f0fdf4">
        <div style="text-align:center;background:white;padding:2rem 3rem;border-radius:1rem;box-shadow:0 4px 24px rgba(0,0,0,.08)">
          <div style="font-size:3rem;margin-bottom:1rem">✅</div>
          <h2 style="color:#111827;margin:0 0 .5rem">Gmail Connected!</h2>
          <p style="color:#6b7280;margin:0 0 1.5rem">Your Gmail account has been authorized successfully.</p>
          <script>setTimeout(()=>window.close(),2000);</script>
          <p style="color:#9ca3af;font-size:.8rem">This window will close automatically…</p>
        </div></body></html>
        """
    except Exception as e:
        return f"<p>OAuth error: {e}</p>", 400


@app.route("/api/gmail/disconnect", methods=["POST"])
def gmail_disconnect():
    if os.path.exists(_GMAIL_TOKEN_FILE):
        os.remove(_GMAIL_TOKEN_FILE)
    return jsonify({"ok": True})


if __name__ == "__main__":
    import logging
    logging.basicConfig(level=logging.DEBUG)
    port = int(os.environ.get("PORT", 8502))
    debug = os.environ.get("RAILWAY_ENVIRONMENT") is None  # debug only locally
    app.run(port=port, debug=debug, use_reloader=debug)
