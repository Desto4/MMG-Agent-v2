"""
Run this once to build the DuckDB database from MMG_Tenant_CRM.xlsx.
Usage:  python build_leads_db.py
The .duckdb file is created next to the Excel file.
"""
import os, sys, re

XLSX = os.path.expanduser("~/Downloads/MMG_Tenant_CRM.xlsx")
if not os.path.isfile(XLSX):
    print(f"ERROR: File not found: {XLSX}")
    print("Edit the XLSX variable at the top of this script to point to your file.")
    sys.exit(1)

DB = os.path.splitext(XLSX)[0] + "_leads.duckdb"

try:
    from openpyxl import load_workbook
except ImportError:
    print("ERROR: openpyxl not installed. Run: pip install openpyxl")
    sys.exit(1)
try:
    import duckdb
except ImportError:
    print("ERROR: duckdb not installed. Run: pip install duckdb")
    sys.exit(1)


def cell_str(v):
    if v is None:
        return ""
    if hasattr(v, "isoformat"):
        return v.isoformat()
    return str(v).strip()


def unique_idents(labels):
    seen = {}
    out = []
    for i, h in enumerate(labels):
        base = h if h.strip() else f"column_{i+1}"
        slug = re.sub(r"[^\w]+", "_", base.strip()).strip("_") or f"col_{i+1}"
        key = slug.lower()
        if key in seen:
            seen[key] += 1
            slug = f"{slug}_{seen[key]}"
        else:
            seen[key] = 0
        out.append('"' + slug.replace('"', '""') + '"')
    return out


print(f"Reading:  {XLSX}")
wb = load_workbook(XLSX, read_only=True, data_only=True)
sheet_names = wb.sheetnames
print(f"Sheets:   {sheet_names}")

# Build schema from first sheet header
first_ws = wb[sheet_names[0]]
first_iter = first_ws.iter_rows(values_only=True)
header_row = next(first_iter)
labels = [str(h).strip() if h and str(h).strip() else f"column_{i+1}" for i, h in enumerate(header_row)]
idents = unique_idents(labels)
all_idents = ['"sheet"'] + idents

if os.path.isfile(DB):
    os.remove(DB)

con = duckdb.connect(database=DB)
cols_sql = '"sheet" VARCHAR, ' + ", ".join(f"{ident} VARCHAR" for ident in idents)
con.execute(f"CREATE TABLE leads_db ({cols_sql})")
placeholders = ", ".join(["?"] * len(all_idents))

total = 0
for sname in sheet_names:
    ws = wb[sname]
    rows_iter = ws.iter_rows(values_only=True)
    next(rows_iter, None)  # skip header
    batch = []
    for row in rows_iter:
        vals = [sname]
        nonempty = False
        for j in range(len(idents)):
            cell = row[j] if j < len(row) else None
            s = cell_str(cell)
            if s:
                nonempty = True
            vals.append(s)
        if nonempty:
            batch.append(tuple(vals))
    if batch:
        con.executemany(f"INSERT INTO leads_db VALUES ({placeholders})", batch)
        total += len(batch)
        print(f"  {sname}: {len(batch)} rows")

wb.close()
con.close()

print(f"\n✓ Done — {total} total rows")
print(f"✓ Database: {DB}")
print(f"\nOpen DBeaver → New Connection → DuckDB → paste this path:\n  {DB}")
